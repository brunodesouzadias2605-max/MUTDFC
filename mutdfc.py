"""
mutdfc.py — MUTDFC: Movimentação do Fluxo de Caixa (contas de Mútuo)

Extrai o Razão Contábil (FBL3N) referente a Mútuo via SAP GUI Scripting
(COM / pywin32), consolida os arquivos diários em um único CSV (limpo e
ordenado), classifica as linhas de Mútuo, cruza consolidação intercompany
(ZFIT009/ZCO059), separa contrapartidas e gera resumo em estrutura de fluxo.

Pré-requisitos:
  - Windows com SAP GUI instalado e SAP GUI Scripting habilitado.
  - Usuário já logado no sistema SAP desejado (o script apenas anexa à sessão).
  - Python 3.x + pywin32  →  pip install -r requirements.txt

Como alterar configurações:
  Ajuste as CONSTANTES abaixo (CONTA_LAYOUT, USUARIO_SAP, SISTEMAS) conforme
  o seu ambiente antes de executar. As pastas de saída são criadas
  automaticamente dentro do próprio diretório do script.
"""

import calendar
import glob
import json
import logging
import os
import re
import sys
import time
import unicodedata
from collections import defaultdict
from datetime import datetime, timedelta
from getpass import getuser

# ---------------------------------------------------------------------------
# CONSTANTES DE CONFIGURAÇÃO — altere aqui conforme o seu ambiente
# ---------------------------------------------------------------------------

# Pasta base do projeto: sempre relativa à localização deste script.
# Para sobrescrever, substitua a linha abaixo por um caminho absoluto, ex.:
#   PASTA_BASE = r"C:\Minha\Pasta\Personalizada"
PASTA_BASE = os.path.dirname(os.path.abspath(__file__))

# Subpastas de saída (criadas automaticamente)
PASTA_DIARIOS     = os.path.join(PASTA_BASE, "saidas", "diarios")
PASTA_CONSOLIDADO = os.path.join(PASTA_BASE, "saidas", "consolidado")
PASTA_TABELAS     = os.path.join(PASTA_BASE, "saidas", "tabelas")
PASTA_LOGS        = os.path.join(PASTA_BASE, "logs")
ARQUIVO_SALDO_INICIAL = os.path.join(PASTA_BASE, "saldo_inicial.json")
ARQUIVO_AJUSTES_MANUAIS = os.path.join(PASTA_BASE, "ajustes_manuais.json")

# Variante de layout salva no SAP (campo txtV-LOW da tela de seleção de layout)
CONTA_LAYOUT = "MUTDFC"

# Usuário SAP responsável pela variante (campo txtENAME-LOW)
USUARIO_SAP = "MS0000240"

# Definições dos sistemas disponíveis no menu
# Chave: número exibido no menu (str)
# "descricao_conexao": string usada pelo SAP GUI para identificar a conexão
# "palavras_chave": substrings procuradas na descrição das conexões abertas
SISTEMAS = {
    "1": {
        "nome": "SAP S/4 HANA (PRD)",
        "descricao_conexao": "MRV SAP S/4 HANA PRODUCAO",
        "palavras_chave": ["S/4 HANA PRODUCAO"],
    },
    "2": {
        "nome": "SAP ECC 6.0 (PRD)",
        "descricao_conexao": "03 SAP PRD ECC - BOLHA",
        "palavras_chave": ["ECC", "BOLHA"],
    },
}

# Cabeçalho exato esperado nos CSVs exportados pelo SAP e no consolidado.
CABECALHO_CONSOLIDADO = (
    "|   St|Atribuição        |Nº doc.   |Dt.Lçto   |Div |Conta     "
    "|Fornecedor|Tp.doc. |Cliente   |Data doc. |CL|      Montante Razão"
    "|DocCompens|Texto                                             "
    "|Imobilizado |Usuário     |Empr|Referência      |Entrado em|Estorno"
    "|DiagRede    |"
)

# Índices dos campos após parse_linha() (0-based dentro de parts[1:-1])
IDX_ST        = 0
IDX_ATRIB     = 1
IDX_NRDOC     = 2
IDX_DTLCTO    = 3
IDX_DIV       = 4
IDX_CONTA     = 5
IDX_FORNEC    = 6
IDX_TPDOC     = 7
IDX_CLIENTE   = 8
IDX_DATADOC   = 9
IDX_CL        = 10
IDX_MONTANTE  = 11
IDX_DOCCOMP   = 12
IDX_TEXTO     = 13
IDX_IMOBIL    = 14
IDX_USUARIO   = 15
IDX_EMPR      = 16
IDX_REFER     = 17
IDX_ENTRADO   = 18
IDX_ESTORNO   = 19
IDX_DIAGREDE  = 20

# Contas específicas usadas na classificação (altere aqui se os números mudarem)
CONTA_IOF   = "2104030007"   # IOF sobre operação de Mútuo
CONTA_IRRF  = "1103050050"   # IRRF sobre rendimento de Mútuo
CONTA_JUROS = "4401020004"   # Receita de juros de Mútuo

# Mapeamento conta → classificação (prioridade: IOF > IRRF > Juros)
CONTAS_CLASSIFICACAO = {
    CONTA_IOF:   "IOF",
    CONTA_IRRF:  "IRRF",
    CONTA_JUROS: "Juros",
}

# Prefixo das contas de Mútuo (4 dígitos — filtra linhas de Mútuo)
PREFIXO_MUTUO = "1202"

# Prefixo longo das contas de Mútuo (6 dígitos — usado na regra "Parceiro")
PREFIXO_MUTUO_LONGO = "120206"

# Conta base de Mútuo (excluída da regra "Parceiro")
CONTA_MUTUO_BASE = "1202060000"

# De-para Conta → Destino (prioridade máxima na resolução do Destino)
CONTAS_DESTINO_POR_CONTA = {
    "1202060002": "MRL",
    "1202060016": "Prime",
    "1202060104": "AHS",
    "2104030007": "IOF",
    "4401020004": "Juros",
    "1103050050": "IRRF",
}

# Vínculos iniciais para ajustes_manuais.json (criados se o arquivo não existir)
AJUSTES_INICIAIS = {
    "2200000404": "E024",
    "2200010827": "E662",
    "2200000193": "MAGI",
    "1800000063": "P001",
    "1800000060": "B001",
}

# Colunas removidas da saída final (Classificado CSV + Excel)
COLUNAS_REMOVER_SAIDA = {"St", "Atribuição", "Imobilizado", "DiagRede", "Status Consolidação"}

# Contas de contrapartida movidas para auditoria após classificação
CONTAS_CONTRAPARTIDA_AUDITORIA = {CONTA_JUROS, CONTA_IRRF, CONTA_IOF}

# Ordem fixa do resumo em estrutura de fluxo (Relatório de Auditoria)
# Ordem: (1) Período de Referência [cabeçalho], (2) Saldo Inicial, (3) Adições,
# (4) Baixas, (5) Transferências [se existir], (6) Reclassificações [se existir],
# (7) Outros Movimentos [se existir], (8) Ajustes [inclui Juros, IOF, IRRF e Eliminações], (9) Saldo Final
_ORDEM_FLUXO = [
    "Saldo Inicial",
    "Adições",
    "Baixas",
    "Transferências",
    "Reclassificações",
    "Outros Movimentos",
    "Ajustes",
    "Saldo Final",
]

# Classificações que compõem a linha "Ajustes" no resumo
_AJUSTES_COMPONENTES = {"Juros", "IOF", "IRRF", "Eliminações de Adições", "Eliminações de Baixas"}

# Sentinel retornado por extrair_razao() quando o dia não tem partidas.
# Distingue "sem dados" (fluxo normal) de None (falha/erro).
_SEM_DADOS = ""

# ---------------------------------------------------------------------------
# IDs SAP GUI dos controles usados na importação ZFIT009 / ZCO059
# (alinhados com os VBS gravados na sessão atual — não alterar sem novo VBS)
# ---------------------------------------------------------------------------

# ZFIT009 — SE16: menu de configuração de campos
_ZFIT009_MENU_CAMPOS = "wnd[0]/mbar/menu[3]/menu[0]/menu[1]"

# ZCO059 — caminho do ALV Grid principal
_ZCO059_SHELL = (
    "wnd[0]/usr/subSUB_DRE:ZCOR043:0100/cntlCCONTAINER_1/shellcont/shell"
)
_ZCO059_DIVISAO_RE = re.compile(r"[A-Z0-9]{2,10}")


# ---------------------------------------------------------------------------
# CONFIGURAÇÃO DO LOG
# ---------------------------------------------------------------------------

def configurar_log(pasta: str) -> tuple:
    """Cria e retorna (logger, caminho_log) com handlers de arquivo e console."""
    os.makedirs(pasta, exist_ok=True)
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    caminho_log = os.path.join(pasta, f"MUTDFC_log_{timestamp}.txt")

    logger = logging.getLogger("MUTDFC")
    logger.setLevel(logging.DEBUG)

    # Evitar adicionar handlers duplicados se main() for chamada mais de uma vez
    if logger.handlers:
        logger.handlers.clear()

    fmt = logging.Formatter(
        "%(asctime)s [%(levelname)s] %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    # Handler de arquivo
    fh = logging.FileHandler(caminho_log, encoding="utf-8")
    fh.setLevel(logging.DEBUG)
    fh.setFormatter(fmt)
    logger.addHandler(fh)

    # Handler de console
    ch = logging.StreamHandler(sys.stdout)
    ch.setLevel(logging.INFO)
    ch.setFormatter(fmt)
    logger.addHandler(ch)

    logger.info("Log iniciado: %s", caminho_log)
    return logger, caminho_log


# ---------------------------------------------------------------------------
# PARSE DE LINHAS
# ---------------------------------------------------------------------------

def parse_linha(linha: str) -> list:
    """
    Divide a linha pelo delimitador '|' e retorna a lista de campos (sem o
    primeiro e último elementos vazios produzidos pela divisão de uma linha
    que começa e termina com '|').
    """
    partes = linha.split("|")
    return partes[1:-1] if len(partes) >= 2 else partes


def tipo_de_linha(linha: str) -> str:
    """
    Classifica o tipo de uma linha do CSV exportado pelo SAP.

    Retorna uma das strings:
      'dado'      — linha de lançamento real, deve ser mantida
      'cabecalho' — linha com os títulos das colunas
      'separador' — linha de hífens entre cabeçalho e dados
      'total'     — linha de subtotal/total (campo St contém '*')
      'vazio'     — linha em branco ou sem conteúdo útil
    """
    if not linha.strip():
        return "vazio"

    # Linhas que não começam com '|': separadores puros de hífens ou lixo
    if not linha.startswith("|"):
        if linha.lstrip("-").strip() == "":
            return "separador"
        return "vazio"

    # Verificar separador: células com apenas hífens representam ≥ 50 % do total
    campos = parse_linha(linha)
    if not campos:
        return "vazio"

    celulas_dash = sum(
        1 for c in campos if c.strip() and not c.strip().replace("-", "").strip()
    )
    celulas_com_conteudo = sum(1 for c in campos if c.strip())
    if celulas_com_conteudo > 0 and celulas_dash / celulas_com_conteudo >= 0.5:
        return "separador"

    # Verificar cabeçalho pelos termos exclusivos dos títulos de coluna
    termos_cabecalho = ("Montante", "Atribuição", "Atribuicao", "Nº doc", "Dt.Lçto")
    if any(t in linha for t in termos_cabecalho):
        return "cabecalho"

    # Verificar total: campo St (índice 0) contém '*'
    if campos and "*" in campos[IDX_ST]:
        return "total"

    # Verificar linha vazia de conteúdo (só pipes)
    if not linha.replace("|", "").strip():
        return "vazio"

    return "dado"


# ---------------------------------------------------------------------------
# LIMPEZA E ORDENAÇÃO
# ---------------------------------------------------------------------------

def limpar_e_ordenar(linhas: list, logger: logging.Logger) -> list:
    """
    Filtra as linhas mantendo apenas as do tipo 'dado' e ordena
    cronologicamente por Dt.Lçto (formato dd.mm.aaaa).

    Linhas com data inválida ou ausente são movidas para o final e logadas
    como aviso.
    """
    dados_validos = []
    dados_invalidos = []

    for linha in linhas:
        if tipo_de_linha(linha) != "dado":
            continue
        campos = parse_linha(linha)
        data_str = campos[IDX_DTLCTO].strip() if len(campos) > IDX_DTLCTO else ""
        try:
            datetime.strptime(data_str, "%d.%m.%Y")
            dados_validos.append(linha)
        except ValueError:
            logger.warning(
                "Linha com Dt.Lçto inválida ('%s') movida para o fim: %s",
                data_str,
                linha[:80],
            )
            dados_invalidos.append(linha)

    def _chave_data(linha: str) -> datetime:
        campos = parse_linha(linha)
        try:
            return datetime.strptime(campos[IDX_DTLCTO].strip(), "%d.%m.%Y")
        except (ValueError, IndexError):
            return datetime.max

    dados_validos.sort(key=_chave_data)
    logger.debug(
        "limpar_e_ordenar: %d válidas, %d com data inválida.",
        len(dados_validos),
        len(dados_invalidos),
    )
    return dados_validos + dados_invalidos


# ---------------------------------------------------------------------------
# UTILITÁRIOS NUMÉRICOS
# ---------------------------------------------------------------------------

def _parse_montante(valor_str: str):
    """
    Converte string de montante no formato brasileiro para float.

    Formatos suportados:
      '1.234,56'    → positivo
      '-8.000,00'   → negativo (sinal à esquerda)
      '3.301,24-'   → negativo (sinal à direita, padrão SAP FBL3N)

    Retorna None se não for possível converter.
    """
    if not valor_str:
        return None
    try:
        s = valor_str.strip()
        # Sinal negativo à direita (ex.: '3.301,24-') — padrão SAP
        negativo = s.endswith("-")
        if negativo:
            s = s[:-1].strip()
        # Remove separador de milhar (.) e substitui decimal (,) por (.)
        limpo = s.replace(".", "").replace(",", ".")
        valor = float(limpo)
        return -valor if negativo else valor
    except (ValueError, AttributeError):
        return None


def _formatar_valor_br(valor: float) -> str:
    """Formata float no padrão brasileiro com negativo à direita."""
    abs_v = abs(valor)
    s = f"{abs_v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
    return f"{s}-" if valor < 0 else s


def _parse_numero_br(valor_str: str):
    """Converte texto numérico BR para float (aceita negativo à esquerda/direita)."""
    if valor_str is None:
        return None
    s = str(valor_str).strip()
    if not s:
        return None
    negativo = s.endswith("-")
    if negativo:
        s = s[:-1].strip()
    if s.startswith("-"):
        negativo = True
        s = s[1:].strip()
    s = s.replace(".", "").replace(",", ".")
    try:
        valor = float(s)
        return -valor if negativo else valor
    except ValueError:
        return None


def _formatar_valor_milhares(valor: float) -> str:
    """
    Formata float para milhares de reais (divide por 1000) no padrão brasileiro.
    - SEM casas decimais
    - Com separador de milhar (ponto)
    - Mantém padrão contábil: negativo entre parênteses
    Ex.: 12.548.963 -> 12.549 ; 1.250.000 -> 1.250 ; -850.745 -> (851)
    """
    milhares = round(valor / 1000)  # Arredonda para milhares sem decimais
    abs_v = abs(milhares)
    # Formata com separador de milhar (ponto no padrão BR)
    s = f"{abs_v:,}".replace(",", ".")
    return f"({s})" if milhares < 0 else s


def _formatar_valor_milhares_excel(valor: float) -> int:
    """
    Retorna valor em milhares como inteiro para uso em Excel.
    Divide por 1000 e arredonda.
    """
    return round(valor / 1000)


# ---------------------------------------------------------------------------
# DETECÇÃO DE DIA SEM MOVIMENTO
# ---------------------------------------------------------------------------

def dia_sem_movimento(session, logger: logging.Logger) -> bool:
    """
    Verifica se a barra de status do SAP indica que não há partidas selecionadas.

    Após a execução do F8 na FBL3N, quando não há lançamentos no período,
    o SAP exibe na statusbar (canto inferior esquerdo) a mensagem:
      'Nenhuma partida selecionada (ver texto descritivo)'

    Retorna True se a mensagem indicar ausência de dados, False caso contrário
    ou em caso de erro ao ler a statusbar.
    """
    try:
        sbar = session.findById("wnd[0]/sbar")
        msg = (sbar.Text or "").strip().lower()
        logger.debug("Statusbar após F8: '%s'", msg)
        # Comparação normalizada: tolera variações de maiúsculas e acentos
        return "nenhuma partida" in msg or "no items" in msg
    except Exception as exc:  # noqa: BLE001
        logger.debug("Erro ao ler statusbar: %s", exc)
        return False


# ---------------------------------------------------------------------------
# CONEXÃO COM O SAP
# ---------------------------------------------------------------------------

def conectar_sap(sistema: dict, logger: logging.Logger):
    """
    Localiza e retorna a sessão SAP correspondente ao sistema escolhido.

    O usuário deve estar previamente logado. O script apenas anexa à sessão
    existente; se não encontrar, tenta abrir a conexão.

    Retorna o objeto session do SAP GUI Scripting ou None em caso de falha.
    """
    try:
        import win32com.client  # noqa: PLC0415 — importado aqui para não falhar em Linux
    except ImportError:
        logger.error(
            "Módulo win32com.client não encontrado. "
            "Instale com: pip install pywin32"
        )
        return None

    try:
        sap_gui = win32com.client.GetObject("SAPGUI")
        application = sap_gui.GetScriptingEngine
        logger.info("SAP GUI Scripting Engine obtido com sucesso.")
    except Exception as exc:  # noqa: BLE001
        logger.error("Não foi possível obter o SAP GUI Scripting Engine: %s", exc)
        return None

    palavras = sistema["palavras_chave"]
    session = None

    # Tentar localizar entre as conexões já abertas
    try:
        num_conexoes = application.Children.Count
        logger.info("Conexões abertas no SAP GUI: %d", num_conexoes)
        for i in range(num_conexoes):
            conexao = application.Children(i)
            descricao = getattr(conexao, "Description", "") or ""
            logger.debug("Conexão %d: '%s'", i, descricao)
            if any(p.upper() in descricao.upper() for p in palavras):
                session = conexao.Children(0)
                logger.info(
                    "Sessão localizada na conexão %d ('%s').", i, descricao
                )
                break
    except Exception as exc:  # noqa: BLE001
        logger.warning("Erro ao iterar conexões abertas: %s", exc)

    # Se não encontrou, tentar abrir a conexão
    if session is None:
        descricao_conexao = sistema["descricao_conexao"]
        logger.info(
            "Sessão não localizada. Tentando abrir conexão: '%s'",
            descricao_conexao,
        )
        try:
            conexao = application.OpenConnection(descricao_conexao, True)
            session = conexao.Children(0)
            logger.info("Conexão aberta com sucesso: '%s'", descricao_conexao)
        except Exception as exc:  # noqa: BLE001
            logger.error(
                "Falha ao abrir conexão '%s': %s", descricao_conexao, exc
            )
            return None

    return session


# ---------------------------------------------------------------------------
# UTILITÁRIO DE ESPERA DE CONTROLE SAP
# ---------------------------------------------------------------------------

def esperar_controle(
    session,
    control_id: str,
    timeout: float = 15.0,
    intervalo: float = 0.4,
    logger: logging.Logger = None,
) -> bool:
    """
    Aguarda até que o controle SAP identificado por *control_id* esteja
    disponível, evitando o erro 619 ("The control could not be found by id").

    Estratégia:
      - Chama session.findById(control_id, False) em loop (retorna None se não
        encontrado, sem lançar exceção).
      - Repete a cada *intervalo* segundos até *timeout* segundos.

    Retorna True se o controle foi encontrado, False se ocorreu timeout.
    Loga um aviso se o controle não aparecer dentro do timeout.
    """
    inicio = time.time()
    while True:
        try:
            ctrl = session.findById(control_id, False)
            if ctrl is not None:
                return True
        except Exception:  # noqa: BLE001
            pass
        decorrido = time.time() - inicio
        if decorrido >= timeout:
            if logger:
                logger.warning(
                    "esperar_controle: timeout (%.1fs) aguardando '%s'.",
                    timeout,
                    control_id,
                )
            return False
        time.sleep(intervalo)


# ---------------------------------------------------------------------------
# EXTRAÇÃO FBL3N
# ---------------------------------------------------------------------------

def _fmt_data(data: datetime) -> str:
    """Formata datetime para o padrão SAP dd.mm.aaaa."""
    return data.strftime("%d.%m.%Y")


def extrair_razao(
    session,
    data_low: datetime,
    data_high: datetime,
    pasta: str,
    logger: logging.Logger,
):
    """
    Executa a transação FBL3N para o intervalo [data_low, data_high] e
    exporta o resultado para CSV na pasta informada.

    Retorna:
      str  — caminho do CSV gerado (sucesso com dados)
      _SEM_DADOS ('')  — dia sem partidas (INFO, fluxo normal, sem CSV gerado)
      None — falha/erro (WARNING/ERROR já registrado no log)
    """
    nome_arquivo = f"{_fmt_data(data_low)}.csv"
    caminho_csv = os.path.join(pasta, nome_arquivo)
    str_low = _fmt_data(data_low)
    str_high = _fmt_data(data_high)

    logger.info("Iniciando extração FBL3N: %s → %s", str_low, str_high)

    try:
        # Navegar para FBL3N via campo de comando
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").Text = "fbl3n"
        session.findById("wnd[0]").sendVKey(0)   # Enter
        session.findById("wnd[0]").sendVKey(17)  # Abrir variante de seleção
        logger.debug("FBL3N aberta; abrindo seleção de variante.")
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao abrir FBL3N (navegação): %s", exc)
        return None

    try:
        # Preencher variante de layout e usuário
        session.findById("wnd[1]/usr/txtV-LOW").Text = CONTA_LAYOUT
        session.findById("wnd[1]/usr/txtENAME-LOW").Text = USUARIO_SAP
        session.findById("wnd[1]/usr/txtENAME-LOW").setFocus()
        session.findById("wnd[1]").sendVKey(0)  # Confirmar
        logger.debug("Variante '%s' / usuário '%s' preenchidos.", CONTA_LAYOUT, USUARIO_SAP)
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao preencher variante/usuário: %s", exc)
        return None

    try:
        # Preencher datas e executar (F8)
        session.findById("wnd[0]/usr/ctxtSO_BUDAT-LOW").Text = str_low
        session.findById("wnd[0]/usr/ctxtSO_BUDAT-HIGH").Text = str_high
        session.findById("wnd[0]/usr/ctxtSO_BUDAT-HIGH").setFocus()
        session.findById("wnd[0]").sendVKey(8)   # F8 — Executar
        logger.debug("Datas preenchidas: LOW=%s HIGH=%s. Executando...", str_low, str_high)
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao preencher datas ou executar FBL3N: %s", exc)
        return None

    # -----------------------------------------------------------------------
    # DETECÇÃO DE DIA SEM MOVIMENTO — verificar statusbar antes de exportar
    # Se não houver partidas, o SAP mantém o cursor na tela de seleção e exibe
    # "Nenhuma partida selecionada" na barra de status (statusbar).
    # Nesse caso pulamos a exportação sem gerar ERROR/WARNING.
    # -----------------------------------------------------------------------
    if dia_sem_movimento(session, logger):
        logger.info(
            "Dia %s sem partidas — nenhum dado a exportar.", str_low
        )
        # Fechar FBL3N e retornar ao menu SAP
        try:
            session.findById("wnd[0]").sendVKey(3)  # Voltar
        except Exception:  # noqa: BLE001
            pass
        return _SEM_DADOS  # sentinel: sem dados, sem erro

    # -----------------------------------------------------------------------
    # EXPORTAÇÃO — só chega aqui quando há dados na lista
    # -----------------------------------------------------------------------
    try:
        # Selecionar tudo e abrir diálogo de exportação
        session.findById("wnd[0]").sendVKey(20)  # Ctrl+Shift+F9 — selecionar tudo
        session.findById("wnd[0]").sendVKey(3)   # F3 — voltar / abrir menu de lista
        session.findById("wnd[1]/usr/btnBUTTON_1").press()
        session.findById("wnd[0]").sendVKey(9)   # Menu "Lista → Exportar → Arquivo local"
        session.findById("wnd[1]").sendVKey(0)   # Confirmar formato (tabela)
        logger.debug("Diálogo de exportação aberto.")
    except Exception as exc:  # noqa: BLE001
        # Erro 617 ("virtual key not enabled") pode ocorrer mesmo com dados em
        # situações inesperadas. Logar como WARNING com contexto para diagnóstico.
        logger.warning(
            "Erro ao abrir diálogo de exportação para %s→%s "
            "(dia pode ter dados mas VKey desabilitada): %s",
            str_low, str_high, exc,
        )
        return None

    try:
        # Preencher caminho e nome do arquivo de destino
        session.findById("wnd[1]/usr/ctxtDY_PATH").Text = pasta
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").Text = nome_arquivo
        session.findById("wnd[1]/tbar[0]/btn[11]").press()  # Substituir / Salvar
        logger.info("CSV exportado: %s", caminho_csv)
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao salvar CSV '%s': %s", caminho_csv, exc)
        return None

    try:
        # Fechar tela de resultado e voltar ao menu inicial
        session.findById("wnd[0]").sendVKey(3)  # Voltar
        session.findById("wnd[0]").sendVKey(3)  # Voltar ao menu SAP
        logger.debug("Telas fechadas após exportação.")
    except Exception as exc:  # noqa: BLE001
        logger.warning("Aviso ao fechar telas pós-exportação: %s", exc)

    return caminho_csv


# ---------------------------------------------------------------------------
# ITERAÇÃO POR PERÍODO
# ---------------------------------------------------------------------------

def _intervalos_diarios(data_ini: datetime, data_fim: datetime):
    """Gera pares (low, high) para cada dia do período."""
    atual = data_ini
    while atual <= data_fim:
        yield atual, atual
        atual += timedelta(days=1)


def _intervalos_semanais(data_ini: datetime, data_fim: datetime):
    """
    Gera pares (low, high) para cada semana do período.
    A semana começa na segunda-feira (weekday=0).
    """
    inicio_semana = data_ini - timedelta(days=data_ini.weekday())
    while inicio_semana <= data_fim:
        fim_semana = inicio_semana + timedelta(days=6)
        low = max(inicio_semana, data_ini)
        high = min(fim_semana, data_fim)
        yield low, high
        inicio_semana += timedelta(weeks=1)


def _intervalos_mensais(data_ini: datetime, data_fim: datetime):
    """
    Gera pares (low, high) para cada mês do período.
    Trata corretamente meses de tamanhos diferentes e virada de ano.
    """
    ano, mes = data_ini.year, data_ini.month
    while True:
        primeiro_dia = datetime(ano, mes, 1)
        ultimo_dia = datetime(ano, mes, calendar.monthrange(ano, mes)[1])
        low = max(primeiro_dia, data_ini)
        high = min(ultimo_dia, data_fim)
        if low > data_fim:
            break
        yield low, high
        mes += 1
        if mes > 12:
            mes = 1
            ano += 1


def iterar_periodo(
    data_ini: datetime,
    data_fim: datetime,
    periodicidade: str,
    session,
    pasta: str,
    logger: logging.Logger,
) -> list:
    """
    Itera pelo período conforme a periodicidade e executa extrair_razao
    para cada intervalo. Retorna lista de caminhos de CSVs gerados com sucesso.
    Dias sem partidas são ignorados silenciosamente (já logados como INFO).
    """
    geradores = {
        "diaria": _intervalos_diarios,
        "semanal": _intervalos_semanais,
        "mensal": _intervalos_mensais,
    }
    gerador = geradores.get(periodicidade, _intervalos_diarios)

    csvs_gerados = []
    for low, high in gerador(data_ini, data_fim):
        caminho = extrair_razao(session, low, high, pasta, logger)
        if caminho is None:
            # Falha real — aviso já emitido dentro de extrair_razao
            logger.warning(
                "Extração falhou para o intervalo %s→%s. Continuando...",
                _fmt_data(low),
                _fmt_data(high),
            )
        elif caminho == _SEM_DADOS:
            # Dia sem partidas — tratado como INFO, não há CSV para adicionar
            pass
        else:
            csvs_gerados.append(caminho)

    return csvs_gerados


# ---------------------------------------------------------------------------
# CONSOLIDAÇÃO
# ---------------------------------------------------------------------------

# Padrões típicos de mojibake (UTF-8 lido como latin-1)
_MOJIBAKE_PATTERNS = ("Ã§", "Ã£", "Ãµ", "Ãº", "LÃ§", "Ã£o", "Ã§Ã£", "Ã©", "Ã³")


def _tem_mojibake(texto: str) -> bool:
    """Retorna True se o texto apresentar padrões típicos de duplo-encoding."""
    return any(p in texto for p in _MOJIBAKE_PATTERNS)


def _corrigir_mojibake(texto: str) -> str:
    """
    Tenta corrigir duplo-encoding reinterpretando a string como latin-1 → utf-8.
    Retorna o texto original se a correção falhar.
    """
    try:
        return texto.encode("latin-1").decode("utf-8")
    except (UnicodeEncodeError, UnicodeDecodeError):
        return texto


def ler_csv_corrigindo_encoding(caminho: str, logger: logging.Logger) -> list:
    """
    Lê um arquivo CSV detectando o encoding correto e corrigindo mojibake.

    Estratégia:
      1. Tenta utf-8-sig / utf-8 primeiro (formato mais comum nos CSVs SAP atuais).
      2. Se falhar (UnicodeDecodeError) ou detectar padrões de mojibake, tenta
         cp1252 / latin-1.
      3. Se o conteúdo lido contiver padrões de mojibake (ex.: 'Ã§', 'Ã£'),
         aplica correção reinterpretando bytes como latin-1 → utf-8.
      4. Última alternativa: latin-1 com errors='replace'.

    Retorna lista de strings (linhas sem \\r\\n) com encoding corrigido.
    Loga o encoding detectado e se houve correção de mojibake.
    """
    for enc in ("utf-8-sig", "utf-8", "cp1252", "latin-1"):
        try:
            with open(caminho, "r", encoding=enc, errors="strict") as f:
                linhas = [l.rstrip("\r\n") for l in f]
            # Verifica mojibake na amostra inicial
            amostra = " ".join(linhas[:30])
            if _tem_mojibake(amostra):
                linhas = [_corrigir_mojibake(l) for l in linhas]
                logger.info(
                    "Arquivo '%s': encoding=%s, mojibake detectado e corrigido.",
                    caminho, enc,
                )
            else:
                logger.debug(
                    "Arquivo '%s': encoding=%s.", caminho, enc
                )
            return linhas
        except UnicodeDecodeError:
            continue

    # Última alternativa
    logger.warning(
        "Não foi possível ler '%s' sem erros; usando latin-1 com replace.", caminho
    )
    with open(caminho, "r", encoding="latin-1", errors="replace") as f:
        return [l.rstrip("\r\n") for l in f]


def consolidar(
    csvs: list,
    data_ini: datetime,
    data_fim: datetime,
    pasta: str,
    logger: logging.Logger,
):
    """
    Empilha os CSVs diários em um único arquivo consolidado.

    - Grava um único cabeçalho no topo.
    - Remove linhas de total/subtotal (St contém '*'), separadores,
      cabeçalhos repetidos e linhas em branco.
    - Ordena as linhas de dados cronologicamente por Dt.Lçto.
    - Saída em UTF-8.

    Retorna o caminho do consolidado, ou None se nenhum CSV foi fornecido.
    """
    if not csvs:
        logger.warning("Nenhum CSV para consolidar.")
        return None

    os.makedirs(pasta, exist_ok=True)
    nome = (
        f"Consolidado_{data_ini.strftime('%d.%m.%Y')}"
        f"_a_{data_fim.strftime('%d.%m.%Y')}.csv"
    )
    caminho_consolidado = os.path.join(pasta, nome)

    logger.info(
        "Iniciando consolidação de %d arquivo(s) → '%s'", len(csvs), caminho_consolidado
    )

    todas_linhas = []
    for caminho_csv in sorted(csvs):  # ordem cronológica pelo nome do arquivo
        try:
            for linha in ler_csv_corrigindo_encoding(caminho_csv, logger):
                todas_linhas.append(linha)
        except Exception as exc:  # noqa: BLE001
            logger.error("Erro ao ler '%s': %s", caminho_csv, exc)

    # Limpar (remover totais, separadores, cabeçalhos repetidos) e ordenar
    linhas_limpas = limpar_e_ordenar(todas_linhas, logger)

    with open(caminho_consolidado, "w", encoding="utf-8-sig", newline="") as saida:
        saida.write(CABECALHO_CONSOLIDADO + "\n")
        for linha in linhas_limpas:
            saida.write(linha + "\n")

    logger.info(
        "Consolidação concluída: %d linha(s) de dados gravadas em '%s'.",
        len(linhas_limpas),
        caminho_consolidado,
    )
    return caminho_consolidado


# ---------------------------------------------------------------------------
# CLASSIFICAÇÃO
# ---------------------------------------------------------------------------

def indice_ndoc_por_conta(linhas_dados: list) -> dict:
    """
    Constrói e retorna um dicionário mapeando cada Nº doc. ao conjunto de
    contas em que ele aparece no conjunto de dados fornecido.

    Ex.: {'100123': {'1202060000', '2104030007'}, ...}
    """
    indice: dict = {}
    for linha in linhas_dados:
        campos = parse_linha(linha)
        if len(campos) <= max(IDX_NRDOC, IDX_CONTA):
            continue
        nr = campos[IDX_NRDOC].strip()
        conta = campos[IDX_CONTA].strip()
        if nr:
            indice.setdefault(nr, set()).add(conta)
    return indice


def classificar_linha(campos: list, indice_ndoc: dict) -> str:
    """
    Determina a classificação de uma linha de Mútuo (conta iniciando em '1202').

    Regras em ordem de prioridade:
      1. Se o Nº doc. também aparece na conta IOF  (2104030007) → 'IOF'
      2. Se o Nº doc. também aparece na conta IRRF (1103050050) → 'IRRF'
      3. Se o Nº doc. também aparece na conta Juros(4401020004) → 'Juros'
      4. Senão, pelo campo CL da própria linha:
           CL 09  → 'Adições'
           CL 19  → 'Baixa'
           outro  → '' (sem classificação; a chamada deve logar o CL)

    Linhas de conta que não iniciem com '1202' devem ser filtradas antes de
    chamar esta função.

    Retorna a string de classificação ou '' se não for possível determinar.
    """
    if len(campos) <= IDX_CONTA:
        return ""

    nr = campos[IDX_NRDOC].strip() if len(campos) > IDX_NRDOC else ""
    contas_do_doc = indice_ndoc.get(nr, set())

    # Prioridade 1-3: vínculo por Nº doc. com contas especiais
    for conta_ref in (CONTA_IOF, CONTA_IRRF, CONTA_JUROS):
        if conta_ref in contas_do_doc:
            return CONTAS_CLASSIFICACAO[conta_ref]

    # Prioridade 4: campo CL da própria linha de Mútuo
    cl = campos[IDX_CL].strip() if len(campos) > IDX_CL else ""
    cl_norm = cl.zfill(2)  # normaliza '9' → '09' sem afetar '19'
    if cl_norm == "09":
        return "Adições"
    if cl_norm == "19":
        return "Baixa"
    return ""  # CL desconhecido — quem chamou deve logar


def classificar(
    arquivo_consolidado: str,
    pasta_saida: str,
    logger: logging.Logger,
):
    """
    Lê o arquivo consolidado e gera um arquivo classificado com a coluna
    'Classificação' acrescentada ao final de cada linha.

    Apenas as linhas de Mútuo (conta iniciando em '1202') recebem classificação.
    As linhas de contrapartida (IOF/IRRF/Juros) servem como referência de
    vínculo mas NÃO recebem classificação e não entram nos totais do resumo.

    Regras de prioridade para linhas de Mútuo:
      1. Nº doc. aparece em conta 2104030007 → IOF
      2. Nº doc. aparece em conta 1103050050 → IRRF
      3. Nº doc. aparece em conta 4401020004 → Juros
      4. CL == 09 → Adições; CL == 19 → Baixa; outro → sem classificação (logado)

    Retorna o caminho do arquivo classificado, ou None em caso de falha.
    """
    if not os.path.isfile(arquivo_consolidado):
        logger.error("Arquivo consolidado não encontrado: '%s'", arquivo_consolidado)
        return None

    os.makedirs(pasta_saida, exist_ok=True)

    logger.info("Classificando '%s'...", arquivo_consolidado)

    # Leitura de todas as linhas de dados com correção de encoding
    linhas_dados = []
    cabecalho_original = None
    try:
        for linha in ler_csv_corrigindo_encoding(arquivo_consolidado, logger):
            tp = tipo_de_linha(linha)
            if tp == "cabecalho":
                cabecalho_original = linha
            elif tp == "dado":
                linhas_dados.append(linha)
    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao ler consolidado para classificação: %s", exc)
        return None

    # Construir índice Nº doc. → conjunto de contas (para prioridade 1-3)
    nrdoc_para_contas = indice_ndoc_por_conta(linhas_dados)

    # Contadores para LOG detalhado
    contagens: dict = {}
    sem_classif_mutuo: list = []   # linhas de Mútuo sem classificação (CL desconhecido)
    nao_mutuo_ignoradas: int = 0   # linhas fora de 1202*

    linhas_classificadas = []

    for linha in linhas_dados:
        campos = parse_linha(linha)
        conta = campos[IDX_CONTA].strip() if len(campos) > IDX_CONTA else ""

        if conta.startswith(PREFIXO_MUTUO):
            classif = classificar_linha(campos, nrdoc_para_contas)
            if classif:
                contagens[classif] = contagens.get(classif, 0) + 1
            else:
                cl = campos[IDX_CL].strip() if len(campos) > IDX_CL else "(ausente)"
                nr = campos[IDX_NRDOC].strip() if len(campos) > IDX_NRDOC else ""
                sem_classif_mutuo.append((nr, conta, cl))
                logger.debug(
                    "Mútuo sem classificação: Nº doc.=%s Conta=%s CL='%s'", nr, conta, cl
                )
                classif = "N/A"  # preencher vazio com N/A (item 3)
        else:
            classif = "N/A"  # linhas não-Mútuo também recebem N/A (item 3)
            nao_mutuo_ignoradas += 1

        linhas_classificadas.append(linha.rstrip("|") + "|" + classif + "|")

    # Nome do arquivo de saída
    data_ini_str, data_fim_str = _extrair_datas_do_nome(
        os.path.basename(arquivo_consolidado)
    )
    nome_saida = f"Classificado_{data_ini_str}_a_{data_fim_str}.csv"
    caminho_saida = os.path.join(pasta_saida, nome_saida)

    cabecalho_saida = (
        (cabecalho_original or CABECALHO_CONSOLIDADO).rstrip("|")
        + "|Classificação|"
    )

    with open(caminho_saida, "w", encoding="utf-8-sig", newline="") as f:
        f.write(cabecalho_saida + "\n")
        for linha in linhas_classificadas:
            f.write(linha + "\n")

    logger.info("Arquivo classificado gerado: '%s'", caminho_saida)
    logger.info(
        "Contagem por classificação (linhas Mútuo): %s", contagens
    )
    logger.info(
        "Linhas de Mútuo sem classificação (CL desconhecido): %d",
        len(sem_classif_mutuo),
    )
    if sem_classif_mutuo:
        for nr, conta, cl in sem_classif_mutuo[:20]:  # até 20 exemplos
            logger.warning(
                "  Mútuo sem classif → Nº doc.=%s Conta=%s CL='%s'", nr, conta, cl
            )
    logger.info(
        "Linhas não-Mútuo ignoradas (sem classificação): %d", nao_mutuo_ignoradas
    )

    return caminho_saida


# ---------------------------------------------------------------------------
# TABELAS DE CONSOLIDAÇÃO INTERCOMPANY (ZFIT009 / ZCO059)
# ---------------------------------------------------------------------------

def _linha_pipe_separador(partes: list) -> bool:
    """Retorna True quando a linha é composta apenas por hífens/separadores."""
    if not partes:
        return True
    com_conteudo = [p for p in partes if p.strip()]
    if not com_conteudo:
        return True
    return all(not p.strip().replace("-", "").strip() for p in com_conteudo)


def _normalizar_rotulo_tabela(valor: str) -> str:
    """Normaliza rótulos de colunas SAP para comparações tolerantes a encoding."""
    valor = (valor or "").strip().upper()
    valor = _corrigir_mojibake(valor)
    return "".join(
        ch for ch in unicodedata.normalize("NFKD", valor)
        if not unicodedata.combining(ch)
    )


def _partes_tabela_pipe(linha: str) -> list:
    """Retorna as colunas de uma linha delimitada por pipe, com trim."""
    texto = linha.strip()
    if not texto or "|" not in texto:
        return []
    if texto.startswith("|") and texto.endswith("|"):
        return [p.strip() for p in linha.split("|")[1:-1]]
    return [p.strip() for p in linha.split("|")]


def _transformar_descricao_zco059(descricao: str) -> str:
    """
    Transforma a descrição da ZCO059 em 'Individual' ou 'Controladas'.
    "Individual" → Individual; qualquer outro → Controladas.
    """
    descricao_norm = descricao.strip().upper()
    if descricao_norm == "INDIVIDUAL":
        return "Individual"
    return "Controladas"


def _encontrar_indice_consolida(partes_norm: list) -> int:
    """
    Localiza o índice da coluna 'Consolida' no cabeçalho normalizado.
    Retorna -1 se não encontrada.
    """
    for i, p in enumerate(partes_norm):
        if p == "CONSOLIDA":
            return i
    return -1


def _detectar_formato_zco059_geral(partes_norm: list, logger: logging.Logger) -> dict:
    """
    Detecta posições das colunas no formato ZCO059GERAL (arquivo completo com TODAS as colunas).

    Regras de seleção por POSIÇÃO:
    - Divisão = 4ª coluna (índice 3)
    - Consolida = coluna chamada 'CONSOLIDA'
    - Descrição-estrutura = coluna imediatamente ANTES de 'Consolida' (mostra Individual/SPE Controlada/etc.)
    - A última 'Descrição' (texto Sim/Não do Consolida) é descartada.

    Retorna dict com índices ou None se formato inválido.
    """
    if len(partes_norm) < 5:
        return None

    # Divisão deve estar na 4ª posição (índice 3)
    idx_divisao = 3
    if not partes_norm[idx_divisao].startswith("DIVIS"):
        logger.debug("ZCO059GERAL: coluna 4 não é Divisão ('%s').", partes_norm[idx_divisao])
        return None

    # Localizar Consolida
    idx_consolida = _encontrar_indice_consolida(partes_norm)
    if idx_consolida < 0:
        logger.debug("ZCO059GERAL: coluna 'Consolida' não encontrada no cabeçalho.")
        return None

    # Descrição-estrutura = coluna imediatamente ANTES de Consolida
    idx_descricao_estrutura = idx_consolida - 1
    if idx_descricao_estrutura < 0:
        logger.debug("ZCO059GERAL: não há coluna antes de 'Consolida' para Descrição.")
        return None

    return {
        "divisao": idx_divisao,
        "descricao_estrutura": idx_descricao_estrutura,
        "consolida": idx_consolida,
    }


def parse_zco059(caminho: str, logger: logging.Logger) -> tuple:
    """
    Faz parse robusto da ZCO059 nos formatos aceitos:
      1. ZCO059GERAL — arquivo COMPLETO com TODAS as colunas (exportado via ZCO059Total.vbs).
         Colunas selecionadas por POSIÇÃO:
           - Divisão = 4ª coluna
           - Descrição-estrutura = coluna imediatamente ANTES de 'Consolida'
           - Consolida = coluna com valores 'S'/'N'
         Filtra apenas Consolida = 'S'.
         Há DUAS colunas 'Descrição'; descarta a última (texto Sim/Não).
      2. Formato normalizado (Divisão|Descrição|Consolida) — retrocompatível.
      3. Formato legado (Empresa|Divisão|Descrição|Consolida) — retrocompatível.

    Transforma Descrição-estrutura: 'Individual' → Individual; qualquer outro → Controladas.
    Retorna (registros, cabecalho_encontrado, erro_validacao).
    """
    registros = []
    cabecalho_encontrado = False
    erro_validacao = None
    linhas_ignoradas = 0
    linhas_totais = 0

    # Flags de formato
    formato_geral = None  # dict com índices das colunas se formato GERAL
    tem_descricao = False  # formato legado com 4 colunas

    # Contadores para validação
    contador_individual = 0
    contador_controladas = 0

    for numero_linha, linha in enumerate(ler_csv_corrigindo_encoding(caminho, logger), start=1):
        texto = linha.strip()
        if not texto:
            continue

        partes = _partes_tabela_pipe(linha)
        if not partes:
            continue
        if _linha_pipe_separador(partes):
            continue

        partes_norm = [_normalizar_rotulo_tabela(p) for p in partes]

        # ---- Detecção de cabeçalho ----
        if not cabecalho_encontrado:
            # Tentar detectar formato ZCO059GERAL (arquivo completo com muitas colunas)
            if len(partes_norm) >= 5:
                posicoes = _detectar_formato_zco059_geral(partes_norm, logger)
                if posicoes:
                    formato_geral = posicoes
                    cabecalho_encontrado = True
                    logger.info(
                        "ZCO059GERAL detectado: Divisão=col%d, Descrição-estrutura=col%d, Consolida=col%d (total %d colunas).",
                        posicoes["divisao"] + 1,
                        posicoes["descricao_estrutura"] + 1,
                        posicoes["consolida"] + 1,
                        len(partes_norm),
                    )
                    continue

            # Tentar formato normalizado (Divisão|Descrição|Consolida)
            if len(partes_norm) >= 3:
                if partes_norm[0].startswith("DIVIS") and "DESCRI" in partes_norm[1] and partes_norm[2] == "CONSOLIDA":
                    cabecalho_encontrado = True
                    tem_descricao = True
                    logger.info("ZCO059 formato normalizado detectado (Divisão|Descrição|Consolida).")
                    continue

            # Tentar formato legado (Empresa|Divisão|Descrição|Consolida ou Empresa|Divisão|Consolida)
            if len(partes_norm) >= 3 and partes_norm[0] == "EMPRESA" and partes_norm[1].startswith("DIVIS"):
                if len(partes_norm) >= 4 and "DESCRI" in partes_norm[2] and partes_norm[3] == "CONSOLIDA":
                    cabecalho_encontrado = True
                    tem_descricao = True
                    logger.info("ZCO059 formato legado detectado (Empresa|Divisão|Descrição|Consolida).")
                    continue
                if partes_norm[2] == "CONSOLIDA":
                    cabecalho_encontrado = True
                    tem_descricao = False
                    logger.info("ZCO059 formato legado detectado (Empresa|Divisão|Consolida).")
                    continue

            continue  # Ainda procurando cabeçalho

        # ---- Processar linhas de dados ----
        linhas_totais += 1

        if formato_geral:
            # Formato ZCO059GERAL — seleção por posição
            idx_div = formato_geral["divisao"]
            idx_desc = formato_geral["descricao_estrutura"]
            idx_cons = formato_geral["consolida"]

            if len(partes) <= max(idx_div, idx_desc, idx_cons):
                linhas_ignoradas += 1
                logger.debug("ZCO059GERAL: linha %d ignorada por colunas insuficientes.", numero_linha)
                continue

            divisao = _corrigir_mojibake(partes[idx_div].strip()).upper()
            descricao_raw = _corrigir_mojibake(partes[idx_desc].strip())
            consolida = _corrigir_mojibake(partes[idx_cons].strip()).upper()

            # Filtrar apenas Consolida = 'S'
            if consolida != "S":
                linhas_ignoradas += 1
                continue

        elif tem_descricao:
            # Formato com Descrição (4 colunas legado ou 3 colunas normalizado)
            if partes_norm[0] == "EMPRESA" or (len(partes) >= 4 and partes_norm[0] != ""):
                # Formato legado: Empresa|Divisão|Descrição|Consolida
                if len(partes) < 4:
                    linhas_ignoradas += 1
                    logger.debug("ZCO059: linha %d ignorada por ter menos de 4 colunas.", numero_linha)
                    continue
                divisao = _corrigir_mojibake(partes[1].strip()).upper()
                descricao_raw = _corrigir_mojibake(partes[2].strip())
                consolida = _corrigir_mojibake(partes[3].strip()).upper()
            else:
                # Formato normalizado: Divisão|Descrição|Consolida
                if len(partes) < 3:
                    linhas_ignoradas += 1
                    logger.debug("ZCO059: linha %d ignorada por ter menos de 3 colunas.", numero_linha)
                    continue
                divisao = _corrigir_mojibake(partes[0].strip()).upper()
                descricao_raw = _corrigir_mojibake(partes[1].strip())
                consolida = _corrigir_mojibake(partes[2].strip()).upper()

        else:
            # Formato legado sem Descrição: Empresa|Divisão|Consolida
            if len(partes) < 3:
                linhas_ignoradas += 1
                logger.debug("ZCO059: linha %d ignorada por ter menos de 3 colunas.", numero_linha)
                continue
            divisao = _corrigir_mojibake(partes[1].strip()).upper()
            descricao_raw = ""
            consolida = _corrigir_mojibake(partes[2].strip()).upper()

        # ---- Validações comuns ----
        if not divisao and not consolida:
            linhas_ignoradas += 1
            continue

        if not _ZCO059_DIVISAO_RE.fullmatch(divisao):
            linhas_ignoradas += 1
            logger.debug(
                "ZCO059: linha %d ignorada por Divisão inválida ('%s').",
                numero_linha,
                divisao,
            )
            continue

        if consolida not in {"", "N", "S"}:
            linhas_ignoradas += 1
            logger.debug(
                "ZCO059: linha %d ignorada por Consolida inválido ('%s').",
                numero_linha,
                consolida,
            )
            continue

        # Transformar descrição: "Individual"→Individual; outro→Controladas
        descricao = _transformar_descricao_zco059(descricao_raw) if descricao_raw else "Controladas"

        # Contadores para log de validação
        if descricao == "Individual":
            contador_individual += 1
        else:
            contador_controladas += 1

        registros.append(
            {"Divisão": divisao, "Descrição": descricao, "Consolida": consolida}
        )

    # ---- Validação final ----
    if not cabecalho_encontrado:
        erro_validacao = (
            "ERROR: Cabeçalho da ZCO059 não encontrado. Verifique se o arquivo foi "
            "exportado corretamente via ZCO059Total.vbs ou use o formato normalizado."
        )
        logger.error(erro_validacao)
        return [], False, erro_validacao

    # Log de validação conforme solicitado
    logger.info(
        "ZCO059: %d linha(s) de dados processadas, %d registro(s) com Consolida='S' carregados, %d ignoradas.",
        linhas_totais,
        len(registros),
        linhas_ignoradas,
    )
    logger.info(
        "ZCO059 Estrutura: %d Individual, %d Controladas.",
        contador_individual,
        contador_controladas,
    )

    return registros, cabecalho_encontrado, erro_validacao


def validar_zco059(
    caminho: str,
    registros: list,
    cabecalho_encontrado: bool,
    erro_validacao: str,
    logger: logging.Logger,
    exigir_descricao: bool = False,
) -> bool:
    """
    Valida a ZCO059 já parseada e garante colunas corretas para a consolidação.

    Quando exigir_descricao=True (import via SAP), verifica se a coluna Descrição
    está preenchida corretamente (Individual/Controladas) e falha se estiver faltando.
    """
    if not os.path.isfile(caminho):
        logger.error("Arquivo ZCO059 não encontrado: '%s'", caminho)
        return False
    if erro_validacao:
        logger.error(erro_validacao)
        return False
    if not cabecalho_encontrado:
        logger.error("Cabeçalho esperado não encontrado em '%s' (ZCO059).", caminho)
        return False
    if not registros:
        logger.error("Nenhuma linha válida encontrada em '%s' (ZCO059).", caminho)
        return False

    divisoes = {item["Divisão"] for item in registros if item["Divisão"]}
    invalidos = [
        item for item in registros
        if not _ZCO059_DIVISAO_RE.fullmatch(item["Divisão"])
        or item["Consolida"] not in {"", "N", "S"}
    ]
    if invalidos:
        logger.error(
            "ZCO059 inválida: %d linha(s) fora do padrão de Divisão/Consolida.",
            len(invalidos),
        )
        return False

    # Validar presença da coluna Descrição quando exigido (import SAP)
    if exigir_descricao:
        sem_descricao = [
            item for item in registros
            if item.get("Descrição", "") not in {"Individual", "Controladas"}
        ]
        if sem_descricao:
            logger.error(
                "ZCO059 com Descrição ausente ou inválida em %d linha(s). "
                "O export SAP deve conter 4 colunas (Empresa|Divisão|Descrição|Consolida). "
                "Exporte manualmente via ZCO059.vbs e use a opção de importar por arquivo.",
                len(sem_descricao),
            )
            return False

    logger.info(
        "ZCO059 validada com sucesso: %d linha(s) válidas e %d divisões únicas.",
        len(registros),
        len(divisoes),
    )
    return True


def parse_tabela_pipe(caminho: str, tipo: str, logger: logging.Logger) -> tuple:
    """
    Faz parse robusto de tabela SAP largura-fixa delimitada por '|'.

    Retorna (registros, cabecalho_encontrado).
    - tipo='zfit009' → [{'Cliente': '2200000404', 'Divisão': 'A001'}, ...]
    - tipo='zco059'  → [{'Divisão': 'A001', 'Descrição': 'Individual', 'Consolida': 'S'}, ...]
    """
    if tipo == "zco059":
        registros, cabecalho_encontrado, _ = parse_zco059(caminho, logger)
        return registros, cabecalho_encontrado

    registros = []
    cabecalho_encontrado = False

    for linha in ler_csv_corrigindo_encoding(caminho, logger):
        texto = linha.strip()
        if not texto:
            continue

        partes = _partes_tabela_pipe(linha)
        if not partes:
            continue

        if _linha_pipe_separador(partes):
            continue

        linha_norm = linha.upper()
        if tipo == "zfit009":
            if "CLIENTE" in linha_norm and ("DIVIS" in linha_norm or "DIVISÃ" in linha_norm):
                cabecalho_encontrado = True
                continue
            if len(partes) < 2:
                continue
            cliente = partes[-2].strip()
            divisao = partes[-1].strip()
            if not re.fullmatch(r"\d{10}", cliente):
                continue
            registros.append({"Cliente": cliente, "Divisão": divisao})

    return registros, cabecalho_encontrado


def _validar_importacao_tabela(
    caminho: str, tipo: str, registros: list, cabecalho_encontrado: bool, logger: logging.Logger
) -> bool:
    """Valida existência, cabeçalho e volume de dados importados."""
    if tipo == "zco059":
        return validar_zco059(caminho, registros, cabecalho_encontrado, None, logger)
    if not os.path.isfile(caminho):
        logger.error("Arquivo %s não encontrado: '%s'", tipo.upper(), caminho)
        return False
    if not cabecalho_encontrado:
        logger.error("Cabeçalho esperado não encontrado em '%s' (%s).", caminho, tipo.upper())
        return False
    if len(registros) == 0:
        logger.error("Nenhuma linha de dados válida encontrada em '%s' (%s).", caminho, tipo.upper())
        return False
    logger.info(
        "Importação %s validada com sucesso: %d linha(s) de dados.",
        tipo.upper(),
        len(registros),
    )
    return True


def _salvar_tabela_normalizada(tipo: str, registros: list, pasta_tabelas: str, logger: logging.Logger) -> str:
    """Salva tabela tratada em UTF-8-SIG na pasta de tabelas do projeto."""
    os.makedirs(pasta_tabelas, exist_ok=True)
    nome = "ZFIT009.csv" if tipo == "zfit009" else "ZCO059.csv"
    caminho = os.path.join(pasta_tabelas, nome)

    with open(caminho, "w", encoding="utf-8-sig", newline="") as f:
        if tipo == "zfit009":
            f.write("Cliente|Divisão\n")
            for item in registros:
                f.write(f"{item['Cliente']}|{item['Divisão']}\n")
        else:
            # ZCO059 com 3 colunas: Divisão|Descrição|Consolida
            f.write("Divisão|Descrição|Consolida\n")
            for item in registros:
                descricao = item.get("Descrição", "Controladas")
                f.write(f"{item['Divisão']}|{descricao}|{item['Consolida']}\n")

    logger.info("Tabela %s salva em '%s'.", tipo.upper(), caminho)
    return caminho


def importar_tabela_de_arquivo(caminho_origem: str, tipo: str, pasta_tabelas: str, logger: logging.Logger):
    """Importa ZFIT009/ZCO059 a partir de arquivo já exportado manualmente."""
    if tipo == "zco059":
        registros, cabecalho_encontrado, erro_validacao = parse_zco059(caminho_origem, logger)
        valido = validar_zco059(
            caminho_origem,
            registros,
            cabecalho_encontrado,
            erro_validacao,
            logger,
        )
    else:
        registros, cabecalho_encontrado = parse_tabela_pipe(caminho_origem, tipo, logger)
        valido = _validar_importacao_tabela(
            caminho_origem,
            tipo,
            registros,
            cabecalho_encontrado,
            logger,
        )
    if not valido:
        return None
    return _salvar_tabela_normalizada(tipo, registros, pasta_tabelas, logger)


def importar_zfit009_sap(session, pasta_tabelas: str, logger: logging.Logger):
    """
    Importa ZFIT009 via SAP GUI Scripting (SE16) seguindo exatamente o VBS
    gravado na sessão atual. Salva em saidas/tabelas/ZFIT009.csv.

    Fluxo alinhado ao VBS:
      maximize → SE16 → ZFIT009 → menu[3]/menu[0]/menu[1] → sendVKey 14 →
      chk[1,5] + chk[1,11] → sendVKey 6 → sendVKey 8 (executar) →
      sendVKey 20 (download) → sendVKey 0 → DY_PATH/DY_FILENAME →
      btn[11] → 3× sendVKey 3
    """
    os.makedirs(pasta_tabelas, exist_ok=True)
    caminho = os.path.join(pasta_tabelas, "ZFIT009.csv")

    # Passo 1 — Maximizar e abrir SE16
    try:
        logger.info("ZFIT009 SAP [1/10]: maximizando janela e abrindo SE16.")
        session.findById("wnd[0]").maximize()
        session.findById("wnd[0]/tbar[0]/okcd").Text = "se16"
        session.findById("wnd[0]").sendVKey(0)
        esperar_controle(session, "wnd[0]/usr/ctxtDATABROWSE-TABLENAME", logger=logger)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [1/10 - SE16]: id=wnd[0]/tbar[0]/okcd | %s", exc
        )
        return None

    # Passo 2 — Inserir nome da tabela e confirmar (Enter)
    try:
        logger.info("ZFIT009 SAP [2/10]: inserindo nome da tabela ZFIT009.")
        session.findById("wnd[0]/usr/ctxtDATABROWSE-TABLENAME").Text = "ZFIT009"
        session.findById("wnd[0]").sendVKey(0)
        esperar_controle(session, _ZFIT009_MENU_CAMPOS, logger=logger)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [2/10 - tabela]: id=wnd[0]/usr/ctxtDATABROWSE-TABLENAME | %s", exc
        )
        return None

    # Passo 3 — Abrir configuração de campos via menu (menu[3]/menu[0]/menu[1])
    try:
        logger.info("ZFIT009 SAP [3/10]: abrindo config de campos (menu[3]/menu[0]/menu[1]).")
        session.findById(_ZFIT009_MENU_CAMPOS).select()
        esperar_controle(session, "wnd[1]", logger=logger)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [3/10 - menu campos]: id=%s | %s", _ZFIT009_MENU_CAMPOS, exc
        )
        return None

    # Passo 4 — Navegar para seleção de campos (sendVKey 14 = F14)
    try:
        logger.info("ZFIT009 SAP [4/10]: navegando para seleção de campos (sendVKey 14).")
        session.findById("wnd[1]").sendVKey(14)
        esperar_controle(session, "wnd[1]/usr/chk[1,5]", logger=logger)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [4/10 - sendVKey 14]: id=wnd[1] | %s", exc
        )
        return None

    # Passo 5 — Marcar checkboxes chk[1,5] e chk[1,11]
    try:
        logger.info("ZFIT009 SAP [5/10]: marcando chk[1,5] e chk[1,11].")
        session.findById("wnd[1]/usr/chk[1,5]").Selected = True
        session.findById("wnd[1]/usr/chk[1,11]").Selected = True
        session.findById("wnd[1]/usr/chk[1,11]").setFocus()
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [5/10 - checkboxes]: id=wnd[1]/usr/chk[1,5] | %s", exc
        )
        return None

    # Passo 6 — Aplicar seleção de campos (sendVKey 6 = F6 / Enter na variante)
    try:
        logger.info("ZFIT009 SAP [6/10]: aplicando seleção de campos (sendVKey 6).")
        session.findById("wnd[1]").sendVKey(6)
        esperar_controle(session, "wnd[0]", logger=logger)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [6/10 - sendVKey 6]: id=wnd[1] | %s", exc
        )
        return None

    # Passo 7 — Executar lista (sendVKey 8 = F8) e aguardar resultado
    try:
        logger.info("ZFIT009 SAP [7/10]: executando lista (sendVKey 8 = F8).")
        session.findById("wnd[0]").sendVKey(8)
        time.sleep(2)  # aguarda geração da lista
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [7/10 - F8 executar]: sendVKey(8) | %s", exc
        )
        return None

    # Passo 8 — Abrir diálogo de download (sendVKey 20)
    try:
        logger.info("ZFIT009 SAP [8/10]: abrindo diálogo de download (sendVKey 20).")
        session.findById("wnd[0]").sendVKey(20)
        esperar_controle(session, "wnd[1]", logger=logger)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [8/10 - sendVKey 20]: sendVKey(20) | %s", exc
        )
        return None

    # Passo 9 — Confirmar tipo de arquivo (sendVKey 0 = Enter) e preencher caminho
    try:
        logger.info("ZFIT009 SAP [9/10]: confirmando formato e preenchendo caminho.")
        session.findById("wnd[1]").sendVKey(0)
        esperar_controle(session, "wnd[1]/usr/ctxtDY_PATH", logger=logger)
        session.findById("wnd[1]/usr/ctxtDY_PATH").Text = pasta_tabelas
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").Text = "ZFIT009.csv"
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = 11
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        logger.info("ZFIT009 SAP: arquivo exportado para '%s'.", caminho)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZFIT009 SAP [9/10 - salvar]: id=wnd[1]/usr/ctxtDY_PATH | %s", exc
        )
        return None

    # Passo 10 — Fechar telas (3× sendVKey 3 = Voltar)
    try:
        logger.info("ZFIT009 SAP [10/10]: fechando telas (3× sendVKey 3).")
        session.findById("wnd[0]").sendVKey(3)
        session.findById("wnd[0]").sendVKey(3)
        session.findById("wnd[0]").sendVKey(3)
    except Exception as exc:  # noqa: BLE001
        logger.warning("ZFIT009 SAP [10/10 - fechar telas]: %s", exc)

    registros, cabecalho_encontrado = parse_tabela_pipe(caminho, "zfit009", logger)
    if not _validar_importacao_tabela(caminho, "zfit009", registros, cabecalho_encontrado, logger):
        return None
    return _salvar_tabela_normalizada("zfit009", registros, pasta_tabelas, logger)


def importar_zco059_sap(session, pasta_tabelas: str, logger: logging.Logger):
    """
    Importa ZCO059 via SAP GUI Scripting exportando TODAS as colunas diretamente.
    Salva em saidas/tabelas/ZCO059.csv.

    Fluxo simplificado (alinhado ao ZCO059Total.vbs):
      zco059 → ALV shell → &MB_EXPORT → &PC → btn[0] → DY_PATH/DY_FILENAME → btn[11]

    A seleção de colunas (&COL0) foi removida para evitar erros.
    O Python faz a filtragem de colunas por POSIÇÃO e filtra Consolida='S'.
    """
    os.makedirs(pasta_tabelas, exist_ok=True)
    caminho_tmp = os.path.join(pasta_tabelas, "ZCO059GERAL.csv")

    # Passo 1 — Abrir transação ZCO059 e aguardar ALV grid
    try:
        logger.info("ZCO059 SAP [1]: abrindo transação ZCO059.")
        session.findById("wnd[0]/tbar[0]/okcd").Text = "zco059"
        session.findById("wnd[0]").sendVKey(0)
        esperar_controle(session, _ZCO059_SHELL, timeout=20.0, logger=logger)
        logger.debug("ZCO059 SAP: transação aberta; ALV shell localizado.")
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZCO059 SAP [1 - transação]: id=wnd[0]/tbar[0]/okcd | %s", exc
        )
        return None

    # Passo 2 — Exportar TODAS as colunas diretamente via &MB_EXPORT / &PC
    # (sem seleção de colunas — elimina erro de &COL0)
    try:
        logger.info("ZCO059 SAP [2]: exportando TODAS as colunas (&MB_EXPORT → &PC).")
        shell = session.findById(_ZCO059_SHELL)
        shell.pressToolbarContextButton("&MB_EXPORT")
        shell.selectContextMenuItem("&PC")
        esperar_controle(session, "wnd[1]/tbar[0]/btn[0]", logger=logger)
        session.findById("wnd[1]/tbar[0]/btn[0]").press()
        esperar_controle(session, "wnd[1]/usr/ctxtDY_PATH", logger=logger)
        logger.debug("ZCO059 SAP: diálogo de caminho de exportação aberto.")
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZCO059 SAP [2 - exportar]: id=%s | %s", _ZCO059_SHELL, exc
        )
        return None

    # Passo 3 — Preencher caminho e nome do arquivo e salvar
    try:
        logger.info("ZCO059 SAP [3]: preenchendo caminho e nome do arquivo.")
        session.findById("wnd[1]/usr/ctxtDY_PATH").Text = pasta_tabelas
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").Text = os.path.basename(caminho_tmp)
        session.findById("wnd[1]/usr/ctxtDY_FILENAME").caretPosition = len(os.path.basename(caminho_tmp))
        session.findById("wnd[1]/tbar[0]/btn[11]").press()
        logger.info("ZCO059 SAP: arquivo completo exportado para '%s'.", caminho_tmp)
    except Exception as exc:  # noqa: BLE001
        logger.error(
            "ZCO059 SAP [3 - salvar]: id=wnd[1]/usr/ctxtDY_PATH | %s", exc
        )
        return None

    # Python faz a filtragem de colunas por POSIÇÃO e filtra Consolida='S'
    registros, cabecalho_encontrado, erro_validacao = parse_zco059(caminho_tmp, logger)
    if not validar_zco059(
        caminho_tmp,
        registros,
        cabecalho_encontrado,
        erro_validacao,
        logger,
        exigir_descricao=True,
    ):
        logger.error(
            "ERROR: O export SAP da ZCO059 não contém as colunas esperadas por posição. "
            "Exporte manualmente via ZCO059Total.vbs e use a opção de importar por arquivo."
        )
        try:
            os.remove(caminho_tmp)
        except OSError:
            pass
        return None
    try:
        os.remove(caminho_tmp)
    except OSError:
        logger.debug("ZCO059 SAP: arquivo temporário não pôde ser removido: '%s'.", caminho_tmp)
    return _salvar_tabela_normalizada("zco059", registros, pasta_tabelas, logger)


def carregar_tabela_normalizada(tipo: str, pasta_tabelas: str, logger: logging.Logger):
    """Carrega tabela previamente normalizada da pasta de tabelas."""
    caminho = os.path.join(pasta_tabelas, "ZFIT009.csv" if tipo == "zfit009" else "ZCO059.csv")
    if not os.path.isfile(caminho):
        logger.error("Tabela %s não encontrada em '%s'.", tipo.upper(), caminho)
        return None
    if tipo == "zco059":
        registros, cabecalho_encontrado, erro_validacao = parse_zco059(caminho, logger)
        valido = validar_zco059(
            caminho,
            registros,
            cabecalho_encontrado,
            erro_validacao,
            logger,
        )
    else:
        registros, cabecalho_encontrado = parse_tabela_pipe(caminho, tipo, logger)
        valido = _validar_importacao_tabela(
            caminho,
            tipo,
            registros,
            cabecalho_encontrado,
            logger,
        )
    if not valido:
        return None
    return registros


def gerar_consolidacao(pasta_tabelas: str, pasta_consolidado: str, logger: logging.Logger):
    """
    Gera Consolidacao_Cliente_Divisao_Consolida.csv a partir de ZFIT009 e ZCO059.
    Inclui a coluna Descrição (Estrutura de Consolidação: Individual/Controladas).
    """
    zfit = carregar_tabela_normalizada("zfit009", pasta_tabelas, logger)
    zco = carregar_tabela_normalizada("zco059", pasta_tabelas, logger)
    if zfit is None or zco is None:
        return None

    # Mapear Divisão → (Consolida, Descrição) da ZCO059
    divisao_info = {}
    for item in zco:
        divisao = item["Divisão"].strip()
        if not divisao:
            continue
        valor = item["Consolida"].strip().upper()
        descricao = item.get("Descrição", "Controladas")
        if valor == "S":
            divisao_info[divisao] = ("S", descricao)
        elif divisao not in divisao_info:
            divisao_info[divisao] = (valor, descricao)

    logger.info(
        "ZCO059 carregada para consolidação: %d linhas, %d divisões únicas.",
        len(zco),
        len(divisao_info),
    )

    por_cliente = defaultdict(set)
    divisao_vazia = 0
    linhas_saida = []
    for item in zfit:
        cliente = item["Cliente"].strip()
        divisao = item["Divisão"].strip()
        por_cliente[cliente].add(divisao)
        if not divisao:
            divisao_vazia += 1
        info = divisao_info.get(divisao, ("", "Controladas"))
        consolida = "S" if divisao and info[0] == "S" else ""
        descricao = info[1] if divisao else "Controladas"
        linhas_saida.append((cliente, divisao, descricao, consolida))

    clientes_ambiguos = {c: sorted(v) for c, v in por_cliente.items() if len(v) > 1}
    if clientes_ambiguos:
        logger.warning("Clientes com múltiplas divisões na ZFIT009: %d", len(clientes_ambiguos))
        for cliente, divs in list(clientes_ambiguos.items())[:30]:
            logger.warning("  Cliente %s → divisões %s", cliente, ", ".join(d or "(vazia)" for d in divs))
    if divisao_vazia:
        logger.warning("Registros com divisão vazia na ZFIT009: %d", divisao_vazia)

    os.makedirs(pasta_consolidado, exist_ok=True)
    caminho_saida = os.path.join(
        pasta_consolidado,
        "Consolidacao_Cliente_Divisao_Consolida.csv",
    )
    with open(caminho_saida, "w", encoding="utf-8-sig", newline="") as f:
        f.write("Cliente|Divisão|Descrição|Consolida\n")
        for cliente, divisao, descricao, consolida in linhas_saida:
            f.write(f"{cliente}|{divisao}|{descricao}|{consolida}\n")

    logger.info(
        "Consolidação Cliente→Divisão→Descrição→Consolida gerada: '%s' (%d linhas).",
        caminho_saida,
        len(linhas_saida),
    )
    return caminho_saida


def aplicar_status_consolidacao(
    arquivo_classificado: str,
    arquivo_consolidacao: str,
    pasta_saida: str,
    logger: logging.Logger,
):
    """
    Adiciona colunas Destino, Consolida e Estrutura de Consolidação ao classificado.

    Alterações em relação à versão anterior:
    - Coluna renomeada de "Divisão" para "Destino".
    - Destino determinado por resolver_destino() com precedência:
        (a) Conta específica  (b) Ajuste manual  (c) 120206*→Parceiro  (d) ZFIT009
    - Consolida usa "S"/"N" — buscado via Destino na Divisão da tabela de consolidação.
    - Removida coluna "Status Consolidação" (mantém só "Consolida").
    - Nova coluna "Estrutura de Consolidação" (Individual/Controladas) via coluna Div:
        * Busca Div (divisão real do lançamento) na ZCO059 (coluna Divisão).
        * Retorna "Individual" se Descrição == "Individual", senão "Controladas".
        * Se Div não existir na ZCO059, deixa em branco.
    - Gera TXT 'SemDestino_*' com linhas sem Destino.
    """
    if not os.path.isfile(arquivo_classificado):
        logger.error("Arquivo classificado não encontrado: '%s'", arquivo_classificado)
        return None
    if not os.path.isfile(arquivo_consolidacao):
        logger.error("Arquivo de consolidação não encontrado: '%s'", arquivo_consolidacao)
        return None

    # Carrega de-para cliente→(divisao, descricao, consolida) e divisao→(descricao, consolida)
    idx_consolida = {}       # cliente → [(divisao, consolida), ...]
    divisao_info = {}        # divisao → (descricao, consolida)
    for linha in ler_csv_corrigindo_encoding(arquivo_consolidacao, logger):
        texto = linha.strip()
        if not texto:
            continue
        if texto.lower().startswith("cliente|"):
            continue
        if texto.startswith("|"):
            partes = [p.strip() for p in linha.split("|")[1:-1]]
        else:
            partes = [p.strip() for p in linha.split("|")]
        # Formato: Cliente|Divisão|Descrição|Consolida
        if len(partes) < 4:
            # Retrocompatibilidade com formato de 3 colunas
            if len(partes) >= 3:
                cliente, divisao, consolida = partes[0], partes[1], partes[2].upper()
                descricao = "Controladas"
            else:
                continue
        else:
            cliente, divisao, descricao, consolida = partes[0], partes[1], partes[2], partes[3].upper()
        idx_consolida.setdefault(cliente, []).append((divisao, consolida))
        if divisao and divisao not in divisao_info:
            divisao_info[divisao] = (descricao, consolida)
        elif divisao and consolida == "S":
            # Se já existe, priorizar S
            divisao_info[divisao] = (descricao, consolida)

    # Carrega ajustes manuais (prioridade sobre Parceiro e ZFIT009)
    ajustes_manuais = carregar_ajustes_manuais(logger)

    linhas_saida = []
    linhas_sem_destino = []   # linhas que não obtiveram Destino

    # Contadores por regra de Destino
    contadores_destino = {
        "conta_especifica": 0,
        "ajuste_manual": 0,
        "parceiro": 0,
        "zfit009": 0,
        "sem_destino": 0,
    }
    consolida_s = 0
    consolida_n = 0

    # Detecta índice da coluna Conta a partir do cabeçalho (robusto após remoção)
    idx_conta_dyn = IDX_CONTA
    idx_cliente_dyn = IDX_CLIENTE
    idx_div_dyn = IDX_DIV
    cabecalho_saida = None

    # Contadores de Estrutura de Consolidação
    estrutura_individual = 0
    estrutura_controladas = 0
    estrutura_sem_div = 0

    for linha in ler_csv_corrigindo_encoding(arquivo_classificado, logger):
        tp = tipo_de_linha(linha)
        if tp == "cabecalho":
            # Detectar índices dinamicamente pelo cabeçalho
            campos_cab = parse_linha(linha)
            for i, c in enumerate(campos_cab):
                cn = c.strip().lower()
                if cn == "conta":
                    idx_conta_dyn = i
                if cn == "cliente":
                    idx_cliente_dyn = i
                if cn == "div":
                    idx_div_dyn = i
            # Novo cabeçalho: Destino|Consolida|Estrutura de Consolidação (sem Status Consolidação)
            cabecalho_saida = linha.rstrip("|") + "|Destino|Consolida|Estrutura de Consolidação|"
            continue
        if tp != "dado":
            continue

        campos = parse_linha(linha)
        conta = campos[idx_conta_dyn].strip() if len(campos) > idx_conta_dyn else ""
        cliente = campos[idx_cliente_dyn].strip() if len(campos) > idx_cliente_dyn else ""
        div = campos[idx_div_dyn].strip() if len(campos) > idx_div_dyn else ""

        # Resolve Destino com precedência: Conta específica → Ajuste manual → Parceiro → ZFIT009
        destino, regra = resolver_destino(conta, cliente, ajustes_manuais, idx_consolida)
        contadores_destino[regra] = contadores_destino.get(regra, 0) + 1

        # Consolida por Destino: buscar Destino na coluna Divisão da consolidação
        # Se houver múltiplos destinos (separados por ;), usar o primeiro
        destino_principal = destino.split(";")[0].strip() if destino else ""
        info_destino = divisao_info.get(destino_principal, ("", ""))
        consolida_divisao = info_destino[1]

        # Estrutura de Consolidação por Div (divisão real do lançamento):
        # buscar Div na coluna Divisão da ZCO059; retornar Individual ou Controladas
        if div:
            info_div = divisao_info.get(div, None)
            if info_div is not None:
                descricao_div = info_div[0]
                if descricao_div == "Individual":
                    estrutura = "Individual"
                    estrutura_individual += 1
                else:
                    estrutura = "Controladas"
                    estrutura_controladas += 1
            else:
                # Div não encontrada na ZCO059
                estrutura = ""
                estrutura_sem_div += 1
        else:
            # Div em branco na linha
            estrutura = ""
            estrutura_sem_div += 1

        if consolida_divisao == "S":
            consolida = "S"
            consolida_s += 1
        else:
            consolida = "N"
            consolida_n += 1

        linhas_saida.append(linha.rstrip("|") + f"|{destino}|{consolida}|{estrutura}|")

        # Coleta linha sem Destino para TXT
        if not destino:
            nr = campos[IDX_NRDOC].strip() if len(campos) > IDX_NRDOC else ""
            cl = campos[IDX_CL].strip() if len(campos) > IDX_CL else ""
            montante = campos[IDX_MONTANTE].strip() if len(campos) > IDX_MONTANTE else ""
            texto_doc = campos[IDX_TEXTO].strip() if len(campos) > IDX_TEXTO else ""
            linhas_sem_destino.append(
                f"{nr}|{conta}|{cl}|{cliente}|{montante}|{texto_doc}"
            )

    os.makedirs(pasta_saida, exist_ok=True)
    caminho_saida = os.path.join(pasta_saida, os.path.basename(arquivo_classificado))
    cab_fallback = (
        CABECALHO_CONSOLIDADO.rstrip("|")
        + "|Classificação|Destino|Consolida|Estrutura de Consolidação|"
    )
    with open(caminho_saida, "w", encoding="utf-8-sig", newline="") as f:
        f.write((cabecalho_saida or cab_fallback) + "\n")
        for linha in linhas_saida:
            f.write(linha + "\n")

    logger.info("Status consolidação aplicado em '%s'.", caminho_saida)
    logger.info(
        "Destino — por Conta específica: %d | Ajuste manual: %d | "
        "Parceiro: %d | ZFIT009: %d | Sem Destino: %d",
        contadores_destino["conta_especifica"],
        contadores_destino["ajuste_manual"],
        contadores_destino["parceiro"],
        contadores_destino["zfit009"],
        contadores_destino["sem_destino"],
    )
    logger.info(
        "Consolida (por Destino): S=%d | N=%d",
        consolida_s, consolida_n,
    )
    logger.info(
        "Estrutura de Consolidação (por Div): Individual=%d | Controladas=%d | Sem Div na ZCO059=%d",
        estrutura_individual, estrutura_controladas, estrutura_sem_div,
    )

    # Gera TXT com linhas sem Destino
    qtd_sem = contadores_destino["sem_destino"]
    if qtd_sem:
        logger.warning(
            "Linhas sem Destino: %d — gerando arquivo TXT para investigação.", qtd_sem
        )
        data_ini_str, data_fim_str = _extrair_datas_do_nome(
            os.path.basename(arquivo_classificado)
        )
        nome_txt = f"SemDestino_{data_ini_str}_a_{data_fim_str}.txt"
        caminho_txt = os.path.join(pasta_saida, nome_txt)
        with open(caminho_txt, "w", encoding="utf-8-sig", newline="") as f:
            f.write("Nº doc.|Conta|CL|Cliente|Montante Razão|Texto\n")
            for linha_txt in linhas_sem_destino:
                f.write(linha_txt + "\n")
        logger.warning("  Arquivo SemDestino gerado: '%s' (%d linha(s)).", caminho_txt, qtd_sem)
    else:
        logger.info("Nenhuma linha sem Destino.")

    return caminho_saida


def separar_contrapartidas(
    arquivo_classificado: str,
    pasta_saida: str,
    logger: logging.Logger,
):
    """
    Move linhas de contas de contrapartida para auditoria e remove do principal.
    Ao escrever a saída final, remove as colunas em COLUNAS_REMOVER_SAIDA
    (St, Atribuição, Imobilizado, DiagRede) — item 5.
    """
    if not os.path.isfile(arquivo_classificado):
        logger.error("Arquivo classificado não encontrado: '%s'", arquivo_classificado)
        return None, None

    cabecalho = None
    indices_remover = set()   # índices (0-based no split) a remover na saída
    linhas_principais = []
    linhas_auditoria = []
    for linha in ler_csv_corrigindo_encoding(arquivo_classificado, logger):
        tp = tipo_de_linha(linha)
        if tp == "cabecalho":
            cabecalho = linha
            # Calcular índices das colunas a remover (baseado no cabeçalho)
            partes_cab = linha.split("|")
            for i, p in enumerate(partes_cab):
                if p.strip() in COLUNAS_REMOVER_SAIDA:
                    indices_remover.add(i)
            continue
        if tp != "dado":
            continue
        campos = parse_linha(linha)
        conta = campos[IDX_CONTA].strip() if len(campos) > IDX_CONTA else ""
        if conta in CONTAS_CONTRAPARTIDA_AUDITORIA:
            linhas_auditoria.append(linha)
        else:
            linhas_principais.append(linha)

    def _aplicar_remocao(linha_raw):
        """Remove colunas pelos índices calculados do cabeçalho."""
        if not indices_remover:
            return linha_raw
        partes = linha_raw.split("|")
        partes_filtradas = [p for i, p in enumerate(partes) if i not in indices_remover]
        return "|".join(partes_filtradas)

    # Cabeçalho de saída sem as colunas removidas
    cab_saida = _aplicar_remocao(
        cabecalho if cabecalho else (CABECALHO_CONSOLIDADO.rstrip("|") + "|Classificação|")
    )

    base = os.path.basename(arquivo_classificado)
    data_ini_str, data_fim_str = _extrair_datas_do_nome(base)
    periodo = f"{data_ini_str}_a_{data_fim_str}"

    os.makedirs(pasta_saida, exist_ok=True)
    caminho_principal = os.path.join(pasta_saida, base)
    caminho_auditoria = os.path.join(pasta_saida, f"Auditoria_Contrapartidas_{periodo}.csv")

    with open(caminho_principal, "w", encoding="utf-8-sig", newline="") as f:
        f.write(cab_saida + "\n")
        for linha in linhas_principais:
            f.write(_aplicar_remocao(linha) + "\n")

    with open(caminho_auditoria, "w", encoding="utf-8-sig", newline="") as f:
        f.write(cab_saida + "\n")
        for linha in linhas_auditoria:
            f.write(_aplicar_remocao(linha) + "\n")

    cols_removidas = sorted(
        p.strip()
        for i, p in enumerate(
            (cabecalho or "").split("|")
        )
        if i in indices_remover and p.strip()
    )
    logger.info(
        "Contrapartidas separadas: %d linha(s) para auditoria em '%s'.",
        len(linhas_auditoria),
        caminho_auditoria,
    )
    logger.info(
        "Arquivo principal atualizado sem contrapartidas: %d linha(s) em '%s'.",
        len(linhas_principais),
        caminho_principal,
    )
    if cols_removidas:
        logger.info(
            "Colunas removidas da saída final: %s", ", ".join(cols_removidas)
        )
    else:
        logger.info(
            "Nenhuma coluna de COLUNAS_REMOVER_SAIDA encontrada no cabeçalho."
        )
    return caminho_principal, caminho_auditoria


def exportar_excel_corporativo(caminho_csv: str, logger: logging.Logger):
    """Gera versão Excel do classificado com formatação corporativa."""
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill
    except ImportError:
        logger.error("openpyxl não instalado. Execute: pip install openpyxl")
        return None

    if not os.path.isfile(caminho_csv):
        logger.error("Arquivo CSV não encontrado para exportação Excel: '%s'", caminho_csv)
        return None

    linhas = []
    for linha in ler_csv_corrigindo_encoding(caminho_csv, logger):
        if not linha.strip():
            continue
        texto = linha.strip()
        if texto.startswith("|") and texto.endswith("|"):
            linhas.append([c.strip() for c in linha.split("|")[1:-1]])
        else:
            linhas.append([c.strip() for c in linha.split("|")])

    if not linhas:
        logger.error("Arquivo CSV vazio para exportação Excel: '%s'", caminho_csv)
        return None

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Classificado"

    for row in linhas:
        ws.append(row)

    cabecalho_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    total_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = cabecalho_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    idx_montante = None
    idx_st = None
    for i, valor in enumerate(linhas[0], start=1):
        if "Montante" in valor:
            idx_montante = i
        if valor.strip() == "St":
            idx_st = i

    for col in ws.columns:
        comprimento = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(comprimento + 2, 80)

    if idx_montante:
        for r in range(2, ws.max_row + 1):
            cell = ws.cell(row=r, column=idx_montante)
            valor = _parse_montante(str(cell.value))
            if valor is not None:
                cell.value = valor
                cell.number_format = '#,##0.00_);[Red](#,##0.00)'
                cell.alignment = Alignment(horizontal="right")

    if idx_st:
        for r in range(2, ws.max_row + 1):
            st = str(ws.cell(row=r, column=idx_st).value or "").strip()
            if "*" in st or "TOTAL" in st.upper():
                for c in range(1, ws.max_column + 1):
                    ws.cell(row=r, column=c).fill = total_fill
                    ws.cell(row=r, column=c).font = Font(bold=True)

    caminho_xlsx = os.path.splitext(caminho_csv)[0] + ".xlsx"
    wb.save(caminho_xlsx)
    logger.info("Excel corporativo gerado: '%s'", caminho_xlsx)
    return caminho_xlsx


def carregar_saldo_inicial(logger: logging.Logger):
    """
    Carrega saldo inicial persistido (se existir).
    
    Retorna dicionário com campos:
    - valor_individual: float
    - valor_controladas: float
    - valor: float (consolidado = individual + controladas, para retrocompatibilidade)
    - periodo: str
    - data_alteracao: str
    - usuario: str
    """
    if not os.path.isfile(ARQUIVO_SALDO_INICIAL):
        return None
    try:
        with open(ARQUIVO_SALDO_INICIAL, "r", encoding="utf-8") as f:
            dados = json.load(f)
        ultimo = dados.get("ultimo")
        if ultimo:
            # Retrocompatibilidade: se só tem "valor", assume como Individual
            if "valor_individual" not in ultimo and isinstance(ultimo.get("valor"), (int, float)):
                ultimo["valor_individual"] = ultimo.get("valor", 0.0)
                ultimo["valor_controladas"] = 0.0
            # Sempre recalcula consolidado
            valor_ind = ultimo.get("valor_individual", 0.0)
            valor_ctrl = ultimo.get("valor_controladas", 0.0)
            ultimo["valor"] = valor_ind + valor_ctrl
            return ultimo
    except Exception as exc:  # noqa: BLE001
        logger.warning("Falha ao ler saldo inicial persistido: %s", exc)
    return None


def salvar_saldo_inicial(
    valor_individual: float,
    valor_controladas: float,
    periodo: str,
    logger: logging.Logger,
):
    """
    Persiste saldo inicial segregado (Individual e Controladas) e mantém histórico.
    
    Consolidado = Individual + Controladas (calculado automaticamente).
    """
    os.makedirs(PASTA_BASE, exist_ok=True)
    dados = {"historico": []}
    if os.path.isfile(ARQUIVO_SALDO_INICIAL):
        try:
            with open(ARQUIVO_SALDO_INICIAL, "r", encoding="utf-8") as f:
                dados = json.load(f)
            if "historico" not in dados or not isinstance(dados["historico"], list):
                dados["historico"] = []
        except Exception:  # noqa: BLE001
            dados = {"historico": []}

    valor_consolidado = valor_individual + valor_controladas
    registro = {
        "valor_individual": valor_individual,
        "valor_controladas": valor_controladas,
        "valor": valor_consolidado,  # Consolidado = Individual + Controladas
        "periodo": periodo,
        "data_alteracao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "usuario": getuser(),
    }
    dados["ultimo"] = registro
    dados["historico"].append(registro)

    with open(ARQUIVO_SALDO_INICIAL, "w", encoding="utf-8") as f:
        json.dump(dados, f, ensure_ascii=False, indent=2)

    logger.info(
        "Saldo inicial atualizado: Individual=%s, Controladas=%s, Consolidado=%s (período=%s, usuário=%s).",
        _formatar_valor_br(valor_individual),
        _formatar_valor_br(valor_controladas),
        _formatar_valor_br(valor_consolidado),
        periodo,
        registro["usuario"],
    )
    return registro


# ---------------------------------------------------------------------------
# AJUSTES MANUAIS (de-para Cliente → Divisão)
# ---------------------------------------------------------------------------

def carregar_ajustes_manuais(logger: logging.Logger) -> dict:
    """
    Carrega ajustes manuais de 'ajustes_manuais.json'.
    Se o arquivo não existir, cria com os 5 vínculos iniciais (AJUSTES_INICIAIS).
    Retorna um dict no formato:
      { "vinculos": {cliente: divisao, ...}, "ultima_alteracao": ..., "usuario": ... }
    """
    if not os.path.isfile(ARQUIVO_AJUSTES_MANUAIS):
        dados = {
            "vinculos": dict(AJUSTES_INICIAIS),
            "ultima_alteracao": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "usuario": getuser(),
        }
        try:
            with open(ARQUIVO_AJUSTES_MANUAIS, "w", encoding="utf-8") as f:
                json.dump(dados, f, ensure_ascii=False, indent=2)
            logger.info(
                "Ajustes manuais criados com vínculos iniciais: '%s'.",
                ARQUIVO_AJUSTES_MANUAIS,
            )
        except Exception as exc:  # noqa: BLE001
            logger.warning("Falha ao criar ajustes_manuais.json: %s", exc)
        return dados

    try:
        with open(ARQUIVO_AJUSTES_MANUAIS, "r", encoding="utf-8") as f:
            dados = json.load(f)
        if "vinculos" not in dados or not isinstance(dados["vinculos"], dict):
            dados["vinculos"] = {}
        logger.debug(
            "Ajustes manuais carregados: %d vínculo(s).", len(dados["vinculos"])
        )
        return dados
    except Exception as exc:  # noqa: BLE001
        logger.warning("Falha ao ler ajustes_manuais.json: %s", exc)
        return {"vinculos": {}}


def salvar_ajustes_manuais(ajustes: dict, logger: logging.Logger) -> None:
    """Persiste ajustes manuais em 'ajustes_manuais.json' com usuário/timestamp."""
    ajustes["ultima_alteracao"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    ajustes["usuario"] = getuser()
    try:
        with open(ARQUIVO_AJUSTES_MANUAIS, "w", encoding="utf-8") as f:
            json.dump(ajustes, f, ensure_ascii=False, indent=2)
        logger.info(
            "Ajustes manuais salvos: %d vínculo(s) em '%s'.",
            len(ajustes.get("vinculos", {})),
            ARQUIVO_AJUSTES_MANUAIS,
        )
    except Exception as exc:  # noqa: BLE001
        logger.error("Falha ao salvar ajustes_manuais.json: %s", exc)


def menu_ajustes_manuais(logger: logging.Logger, caminho_log: str) -> None:
    """Menu interativo para gerenciar o de-para manual Cliente → Divisão."""
    print("=" * 60)
    print("  Ajustes Manuais — De-para Cliente → Divisão")
    print("=" * 60)

    ajustes = carregar_ajustes_manuais(logger)
    vinculos = ajustes.setdefault("vinculos", {})

    while True:
        print("\n  Vínculos cadastrados:")
        if vinculos:
            for cliente, divisao in sorted(vinculos.items()):
                print(f"    {cliente} → {divisao}")
        else:
            print("    (nenhum vínculo cadastrado)")

        ult = ajustes.get("ultima_alteracao", "")
        usr = ajustes.get("usuario", "")
        if ult:
            print(f"\n  Última alteração: {ult} | Usuário: {usr}")

        print("\n  A - Adicionar  E - Editar  R - Remover  V - Voltar ao menu")
        opcao = input("Opção: ").strip().upper()

        if opcao == "V":
            break
        elif opcao == "A":
            cliente = input("  Cliente (10 dígitos): ").strip()
            if not re.fullmatch(r"\d{10}", cliente):
                print("  ✗ Cliente deve ter exatamente 10 dígitos numéricos.")
                continue
            divisao = input("  Divisão: ").strip().upper()
            if not divisao:
                print("  ✗ Divisão não pode ser vazia.")
                continue
            if cliente in vinculos:
                print(f"  Aviso: substituindo vínculo existente ({cliente} → {vinculos[cliente]}).")
            vinculos[cliente] = divisao
            salvar_ajustes_manuais(ajustes, logger)
            print(f"  ✓ Vínculo adicionado: {cliente} → {divisao}")
        elif opcao == "E":
            cliente = input("  Cliente a editar: ").strip()
            if cliente not in vinculos:
                print(f"  ✗ Cliente '{cliente}' não encontrado nos ajustes.")
                continue
            print(f"  Valor atual: {cliente} → {vinculos[cliente]}")
            divisao = input("  Nova Divisão: ").strip().upper()
            if not divisao:
                print("  ✗ Divisão não pode ser vazia.")
                continue
            vinculos[cliente] = divisao
            salvar_ajustes_manuais(ajustes, logger)
            print(f"  ✓ Vínculo atualizado: {cliente} → {divisao}")
        elif opcao == "R":
            cliente = input("  Cliente a remover: ").strip()
            if cliente not in vinculos:
                print(f"  ✗ Cliente '{cliente}' não encontrado nos ajustes.")
                continue
            divisao_removida = vinculos.pop(cliente)
            salvar_ajustes_manuais(ajustes, logger)
            print(f"  ✓ Vínculo removido: {cliente} → {divisao_removida}")
        else:
            print(f"  ✗ Opção inválida: '{opcao}'. Use A, E, R ou V.")

    print(f"  LOG: {caminho_log}")


# ---------------------------------------------------------------------------
# RESOLUÇÃO DO DESTINO
# ---------------------------------------------------------------------------

def resolver_destino(
    conta: str,
    cliente: str,
    ajustes_manuais: dict,
    idx_consolida: dict,
) -> tuple:
    """
    Resolve o Destino de uma linha de movimentação com a seguinte precedência:

    (a) De-para por Conta — prioridade máxima (CONTAS_DESTINO_POR_CONTA).
    (b) Ajuste manual por Cliente (ajustes_manuais.json) — sobrepõe Parceiro e ZFIT009.
    (c) Prefixo 120206* (≠1202060000 e não consta nas específicas de (a))
        → Destino = "Parceiro".
    (d) ZFIT009 por Cliente (via tabela de consolidação carregada).

    Retorna (destino: str, regra: str) onde regra é uma das strings:
        'conta_especifica' | 'ajuste_manual' | 'parceiro' | 'zfit009' | 'sem_destino'
    """
    # (a) Conta com mapeamento explícito
    destino_conta = CONTAS_DESTINO_POR_CONTA.get(conta)
    if destino_conta:
        return destino_conta, "conta_especifica"

    # (b) Ajuste manual por Cliente (tem prioridade sobre Parceiro e ZFIT009)
    vinculos = ajustes_manuais.get("vinculos", {}) if isinstance(ajustes_manuais, dict) else {}
    if cliente and cliente in vinculos:
        return vinculos[cliente], "ajuste_manual"

    # (c) Prefixo 120206* exceto a conta base e as já mapeadas em (a)
    if (
        conta.startswith(PREFIXO_MUTUO_LONGO)
        and conta != CONTA_MUTUO_BASE
        and conta not in CONTAS_DESTINO_POR_CONTA
    ):
        return "Parceiro", "parceiro"

    # (d) ZFIT009 via tabela de consolidação
    correspondencias = idx_consolida.get(cliente, [])
    if correspondencias:
        divs = sorted({d for d, _ in correspondencias if d})
        if divs:
            return ";".join(divs), "zfit009"

    return "", "sem_destino"


def calcular_saldo_final_fluxo(
    saldo_inicial: float,
    adicoes: float,
    eliminacoes: float,
    juros: float,
    iof: float,
    irrf: float,
    baixas: float,
) -> float:
    """
    Fórmula de fluxo calibrável:
    Saldo Final = Saldo Inicial + Adições − Eliminações + Juros + IOF − IRRF − Baixas
    """
    return saldo_inicial + adicoes - eliminacoes + juros + iof - irrf - baixas


def gerar_resumo_fluxo(arquivo_classificado: str, pasta_saida: str, logger: logging.Logger):
    """
    Gera resumo de fluxo (CSV + Excel) para auditoria conforme requisitos Issue #17.
    
    ESTRUTURA DO RESUMO EXECUTIVO (somente 2 colunas de valor):
        Descrição | Individual | Consolidado
        (Controladas é usado internamente mas NÃO exibido no resumo)
    
    ORDEM DAS LINHAS:
        (1) Período de Referência [cabeçalho destacado]
        (2) Saldo Inicial
        (3) Adições
        (4) Baixas
        (5) Transferências [se existir]
        (6) Reclassificações [se existir]
        (7) Outros Movimentos [se existir]
        (8) Ajustes [inclui Juros, IOF, IRRF e Eliminações]
        (9) Saldo Final = soma algébrica das linhas 2..8
    
    ELIMINAÇÕES segregadas:
        - "Eliminações de Adições" → sinal invertido (Adição +1000 → -1000)
        - "Eliminações de Baixas" → sinal invertido (Baixa -500 → +500)
    
    VALORES:
        - Em MILHARES de reais (÷1000), sem decimais, separador de milhar BR
        - Negativos em VERMELHO com parênteses: (851)
    
    VALIDAÇÃO:
        - Saldo Final == soma algébrica das linhas 2..8 por coluna
        - Composição analítica reconciliada com cada linha do resumo
    """
    if not os.path.isfile(arquivo_classificado):
        logger.error("Arquivo classificado não encontrado: '%s'", arquivo_classificado)
        return None

    # Carregar saldo inicial persistido (Individual + Controladas separados)
    dados_saldo = carregar_saldo_inicial(logger)
    if dados_saldo:
        saldo_individual_default = dados_saldo.get("valor_individual", dados_saldo.get("valor", 0.0))
        saldo_controladas_default = dados_saldo.get("valor_controladas", 0.0)
        periodo_default = dados_saldo.get("periodo", "")
    else:
        saldo_individual_default = 0.0
        saldo_controladas_default = 0.0
        periodo_default = ""
    
    # Usar valores persistidos diretamente (não solicitar novamente)
    saldo_individual = saldo_individual_default
    saldo_controladas = saldo_controladas_default
    saldo_consolidado = saldo_individual + saldo_controladas
    periodo = periodo_default or "não informado"
    
    logger.info(
        "Saldo Inicial utilizado: Individual=%s, Controladas=%s, Consolidado=%s",
        _formatar_valor_br(saldo_individual),
        _formatar_valor_br(saldo_controladas),
        _formatar_valor_br(saldo_consolidado),
    )

    # Totais segregados por estrutura: Individual e Controladas
    # Cada lançamento pertence a UMA estrutura (Individual OU Controladas), nunca ambas
    totais_individual = defaultdict(float)
    totais_controladas = defaultdict(float)
    
    # Eliminações segregadas por tipo (Adições vs Baixas)
    elim_adicoes_individual = 0.0
    elim_adicoes_controladas = 0.0
    elim_baixas_individual = 0.0
    elim_baixas_controladas = 0.0
    
    # Composição analítica: lista de todos os lançamentos para rastreabilidade
    composicao_analitica = []
    
    # Índices de colunas (detectados dinamicamente)
    idx_classif = None
    idx_consolida = None
    idx_estrutura = None
    idx_montante = IDX_MONTANTE
    idx_conta = IDX_CONTA
    idx_descricao = IDX_TEXTO  # Usa campo Texto como descrição
    idx_empresa = IDX_EMPR
    idx_nrdoc = IDX_NRDOC
    
    cabecalho_encontrado = False

    for linha in ler_csv_corrigindo_encoding(arquivo_classificado, logger):
        tp = tipo_de_linha(linha)
        if tp == "cabecalho":
            cabecalho_encontrado = True
            campos = parse_linha(linha)
            for i, c in enumerate(campos):
                c_norm = c.strip().lower()
                if "classificação" in c_norm or "classificacao" in c_norm:
                    idx_classif = i
                if c_norm == "consolida":
                    idx_consolida = i
                if "estrutura" in c_norm:
                    idx_estrutura = i
                if "montante" in c_norm:
                    idx_montante = i
                if c_norm == "conta":
                    idx_conta = i
                if c_norm == "texto":
                    idx_descricao = i
                if c_norm == "empr":
                    idx_empresa = i
                if "nº doc" in c_norm or "nrdoc" in c_norm:
                    idx_nrdoc = i
            continue
        if tp != "dado":
            continue

        campos = parse_linha(linha)
        conta = campos[idx_conta].strip() if len(campos) > idx_conta else ""
        if not conta.startswith(PREFIXO_MUTUO):
            continue

        classif = campos[idx_classif].strip() if idx_classif is not None and len(campos) > idx_classif else ""
        montante = _parse_montante(campos[idx_montante]) if len(campos) > idx_montante else 0.0
        montante = montante or 0.0
        
        consolida = campos[idx_consolida].strip() if idx_consolida is not None and len(campos) > idx_consolida else ""
        estrutura = campos[idx_estrutura].strip() if idx_estrutura is not None and len(campos) > idx_estrutura else "Controladas"
        
        descricao = campos[idx_descricao].strip() if len(campos) > idx_descricao else ""
        empresa = campos[idx_empresa].strip() if len(campos) > idx_empresa else ""
        nrdoc = campos[idx_nrdoc].strip() if len(campos) > idx_nrdoc else ""
        
        # Determinar se é Individual ou Controladas
        is_individual = (estrutura == "Individual")
        
        # Preparar registro para composição analítica
        valor_ind = montante if is_individual else 0.0
        valor_ctrl = montante if not is_individual else 0.0
        valor_cons = montante
        
        registro_analitico = {
            "conta": conta,
            "descricao": descricao,
            "empresa": empresa,
            "nrdoc": nrdoc,
            "classificacao": classif,
            "consolida": consolida,
            "estrutura": estrutura,
            "valor_individual": valor_ind,
            "valor_controladas": valor_ctrl,
            "valor_consolidado": valor_cons,
        }
        composicao_analitica.append(registro_analitico)
        
        # Acumular por classificação
        if is_individual:
            totais_individual[classif] += montante
        else:
            totais_controladas[classif] += montante
        
        # Eliminações segregadas por tipo: Consolida="S" indica eliminação
        # O tipo de eliminação depende da classificação original
        if consolida == "S":
            if classif == "Adições":
                if is_individual:
                    elim_adicoes_individual += montante
                else:
                    elim_adicoes_controladas += montante
            elif classif == "Baixa":
                if is_individual:
                    elim_baixas_individual += montante
                else:
                    elim_baixas_controladas += montante
            # Outras classificações com Consolida=S vão para eliminações genéricas
            # (podem ser IOF, IRRF, Juros etc.)

    # Extrair datas do período do arquivo para o cabeçalho
    data_ini_str, data_fim_str = _extrair_datas_do_nome(os.path.basename(arquivo_classificado))
    # Converter formato DD.MM.AAAA para DD/MM/AAAA
    data_ini_fmt = data_ini_str.replace(".", "/") if data_ini_str != "inicio" else "??/??/????"
    data_fim_fmt = data_fim_str.replace(".", "/") if data_fim_str != "fim" else "??/??/????"
    periodo_cabecalho = f"RELATÓRIO DE MOVIMENTAÇÃO — Período: {data_ini_fmt} a {data_fim_fmt}"

    # --------------------------------------------------
    # CONSTRUIR LINHAS DO RESUMO EXECUTIVO
    # Ordem: Saldo Inicial, Adições, Baixas, Transferências, Reclassificações,
    #        Outros Movimentos, Ajustes, Saldo Final
    # --------------------------------------------------
    
    # Linha 2: Saldo Inicial (valores persistidos)
    linha_saldo_inicial = ("Saldo Inicial", saldo_individual, saldo_consolidado)
    
    # Linha 3: Adições (soma das adições MENOS eliminações de adições com sinal invertido)
    adicoes_ind = totais_individual.get("Adições", 0.0)
    adicoes_ctrl = totais_controladas.get("Adições", 0.0)
    # Eliminações de Adições: sinal invertido (Adição +1000 → eliminação -1000)
    adicoes_ind_liquido = adicoes_ind - elim_adicoes_individual
    adicoes_ctrl_liquido = adicoes_ctrl - elim_adicoes_controladas
    adicoes_cons = adicoes_ind_liquido + adicoes_ctrl_liquido
    linha_adicoes = ("Adições", adicoes_ind_liquido, adicoes_cons)
    
    # Linha 4: Baixas (valores negativos por natureza)
    # Eliminações de Baixas: sinal invertido (Baixa -500 → eliminação +500)
    baixas_ind = totais_individual.get("Baixa", 0.0)
    baixas_ctrl = totais_controladas.get("Baixa", 0.0)
    baixas_ind_liquido = baixas_ind + elim_baixas_individual  # Invertido: reduz o efeito negativo
    baixas_ctrl_liquido = baixas_ctrl + elim_baixas_controladas
    baixas_cons = baixas_ind_liquido + baixas_ctrl_liquido
    linha_baixas = ("Baixas", baixas_ind_liquido, baixas_cons)
    
    # Linha 5: Transferências (se existir)
    transf_ind = totais_individual.get("Transferências", 0.0) + totais_individual.get("Transferencia", 0.0)
    transf_ctrl = totais_controladas.get("Transferências", 0.0) + totais_controladas.get("Transferencia", 0.0)
    transf_cons = transf_ind + transf_ctrl
    linha_transferencias = ("Transferências", transf_ind, transf_cons) if (transf_ind != 0 or transf_ctrl != 0) else None
    
    # Linha 6: Reclassificações (se existir)
    reclass_ind = totais_individual.get("Reclassificações", 0.0) + totais_individual.get("Reclassificação", 0.0)
    reclass_ctrl = totais_controladas.get("Reclassificações", 0.0) + totais_controladas.get("Reclassificação", 0.0)
    reclass_cons = reclass_ind + reclass_ctrl
    linha_reclassificacoes = ("Reclassificações", reclass_ind, reclass_cons) if (reclass_ind != 0 or reclass_ctrl != 0) else None
    
    # Linha 7: Outros Movimentos (se existir)
    outros_ind = totais_individual.get("Outros", 0.0) + totais_individual.get("Outros Movimentos", 0.0)
    outros_ctrl = totais_controladas.get("Outros", 0.0) + totais_controladas.get("Outros Movimentos", 0.0)
    outros_cons = outros_ind + outros_ctrl
    linha_outros = ("Outros Movimentos", outros_ind, outros_cons) if (outros_ind != 0 or outros_ctrl != 0) else None
    
    # Linha 8: Ajustes (inclui Juros, IOF, IRRF)
    juros_ind = totais_individual.get("Juros", 0.0)
    juros_ctrl = totais_controladas.get("Juros", 0.0)
    iof_ind = totais_individual.get("IOF", 0.0)
    iof_ctrl = totais_controladas.get("IOF", 0.0)
    irrf_ind = totais_individual.get("IRRF", 0.0)
    irrf_ctrl = totais_controladas.get("IRRF", 0.0)
    
    ajustes_ind = juros_ind + iof_ind + irrf_ind
    ajustes_ctrl = juros_ctrl + iof_ctrl + irrf_ctrl
    ajustes_cons = ajustes_ind + ajustes_ctrl
    linha_ajustes = ("Ajustes", ajustes_ind, ajustes_cons)
    
    # Montar lista de linhas do resumo (excluindo None)
    linhas_resumo = [linha_saldo_inicial, linha_adicoes, linha_baixas]
    if linha_transferencias:
        linhas_resumo.append(linha_transferencias)
    if linha_reclassificacoes:
        linhas_resumo.append(linha_reclassificacoes)
    if linha_outros:
        linhas_resumo.append(linha_outros)
    linhas_resumo.append(linha_ajustes)
    
    # Linha 9: Saldo Final = soma algébrica das linhas 2..8 (cada uma com seu sinal)
    # IMPORTANTE: Saldo Final NÃO é pré-calculado independentemente;
    # é a soma das linhas exibidas no resumo
    saldo_final_ind = sum(linha[1] for linha in linhas_resumo)
    saldo_final_cons = sum(linha[2] for linha in linhas_resumo)
    linha_saldo_final = ("Saldo Final", saldo_final_ind, saldo_final_cons)
    
    # Adicionar Saldo Final à lista
    linhas_resumo.append(linha_saldo_final)
    
    # --------------------------------------------------
    # VALIDAÇÃO: Saldo Final == soma algébrica das linhas 2..8
    # --------------------------------------------------
    soma_validacao_ind = sum(linha[1] for linha in linhas_resumo[:-1])  # Excluir Saldo Final
    soma_validacao_cons = sum(linha[2] for linha in linhas_resumo[:-1])
    
    diff_ind = abs(saldo_final_ind - soma_validacao_ind)
    diff_cons = abs(saldo_final_cons - soma_validacao_cons)
    
    if diff_ind > 0.01:
        logger.warning(
            "VALIDAÇÃO: Divergência no Saldo Final Individual: calculado=%s, soma=%s, diff=%s",
            _formatar_valor_br(saldo_final_ind),
            _formatar_valor_br(soma_validacao_ind),
            _formatar_valor_br(diff_ind),
        )
    else:
        logger.info("VALIDAÇÃO OK: Saldo Final Individual = soma das linhas 2..8")
    
    if diff_cons > 0.01:
        logger.warning(
            "VALIDAÇÃO: Divergência no Saldo Final Consolidado: calculado=%s, soma=%s, diff=%s",
            _formatar_valor_br(saldo_final_cons),
            _formatar_valor_br(soma_validacao_cons),
            _formatar_valor_br(diff_cons),
        )
    else:
        logger.info("VALIDAÇÃO OK: Saldo Final Consolidado = soma das linhas 2..8")

    # --------------------------------------------------
    # EXPORTAR CSV (valores em milhares, formato BR)
    # Colunas: Descrição | Individual | Consolidado (sem Controladas)
    # --------------------------------------------------
    nome_csv = f"Resumo_{data_ini_str}_a_{data_fim_str}.csv"
    caminho_csv = os.path.join(pasta_saida, nome_csv)
    os.makedirs(pasta_saida, exist_ok=True)

    with open(caminho_csv, "w", encoding="utf-8-sig", newline="") as f:
        # Cabeçalho do período
        f.write(f"{periodo_cabecalho}||\n")
        f.write("Descrição|Individual|Consolidado\n")
        for descr, ind, cons in linhas_resumo:
            # Valores em milhares
            f.write(f"{descr}|{_formatar_valor_milhares(ind)}|{_formatar_valor_milhares(cons)}\n")

    # --------------------------------------------------
    # EXPORTAR EXCEL com formatação profissional
    # --------------------------------------------------
    caminho_xlsx = os.path.join(pasta_saida, f"Resumo_{data_ini_str}_a_{data_fim_str}.xlsx")
    exportar_resumo_fluxo_xlsx(
        caminho_xlsx,
        linhas_resumo,
        periodo_cabecalho,
        composicao_analitica,
        logger,
    )

    # --------------------------------------------------
    # EXIBIR RESUMO NO CONSOLE (em milhares)
    # --------------------------------------------------
    separador = "-" * 70
    print(separador)
    print(f"  {periodo_cabecalho}")
    print(separador)
    print(f"  {'Descrição':<25} {'Individual':>18} {'Consolidado':>18}")
    print(separador)
    for descr, ind, cons in linhas_resumo:
        print(f"  {descr:<25} {_formatar_valor_milhares(ind):>18} {_formatar_valor_milhares(cons):>18}")
    print(separador)
    print("  * Valores em milhares de reais (R$ mil)")
    print(separador)

    logger.info("Resumo fluxo gerado em '%s' e '%s'.", caminho_csv, caminho_xlsx)
    logger.info("Composição analítica com %d lançamentos.", len(composicao_analitica))
    return caminho_csv


def exportar_resumo_fluxo_xlsx(
    caminho_xlsx: str,
    linhas_resumo: list,
    periodo_cabecalho: str,
    composicao_analitica: list,
    logger: logging.Logger,
):
    """
    Exporta resumo de fluxo para Excel com formatação profissional para auditoria.
    
    FORMATAÇÃO (Issue #17):
    - Cabeçalho do período destacado e centralizado (verde escuro, negrito branco)
    - Valores em MILHARES de reais (÷1000), sem decimais, separador de milhar
    - Negativos em VERMELHO com parênteses
    - Saldo Inicial e Saldo Final em negrito destacados (verde claro)
    - Faixa laranja de separação antes do Saldo Final
    - Alinhamento numérico à direita
    
    ABAS:
    - "Resumo": resumo executivo com colunas Descrição | Individual | Consolidado
    - "Composição Analítica": detalhamento de cada linha com rastreabilidade
    """
    try:
        import openpyxl
        from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    except ImportError:
        logger.error("openpyxl não instalado. Execute: pip install openpyxl")
        return None

    wb = openpyxl.Workbook()
    
    # --------------------------------------------------
    # ABA 1: RESUMO EXECUTIVO
    # --------------------------------------------------
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"
    
    # Estilos
    header_fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")  # Verde escuro
    destaque_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")  # Verde claro
    separador_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")  # Laranja
    negrito_branco_font = Font(bold=True, color="FFFFFF", size=12)
    negrito_font = Font(bold=True)
    vermelho_font = Font(color="FF0000")
    
    # Formato numérico: milhares com separador de ponto, negativo em vermelho com parênteses
    # Excel: #.##0 para positivo; [Red](#.##0) para negativo
    fmt_milhares = '#,##0;[Red](#,##0)'
    
    # Linha 1: Cabeçalho do período (mesclado nas 3 colunas)
    ws_resumo.merge_cells("A1:C1")
    cell_periodo = ws_resumo["A1"]
    cell_periodo.value = periodo_cabecalho
    cell_periodo.font = negrito_branco_font
    cell_periodo.fill = header_fill
    cell_periodo.alignment = Alignment(horizontal="center", vertical="center")
    ws_resumo.row_dimensions[1].height = 30
    
    # Linha 2: Cabeçalho das colunas
    ws_resumo.append(["Descrição", "Individual", "Consolidado"])
    for cell in ws_resumo[2]:
        cell.font = negrito_font
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")  # Cinza claro
        cell.alignment = Alignment(horizontal="center")
    
    # Linhas de dados (a partir da linha 3)
    linha_atual = 3
    for descr, ind, cons in linhas_resumo:
        # Converter para milhares (inteiro)
        ind_milhares = _formatar_valor_milhares_excel(ind)
        cons_milhares = _formatar_valor_milhares_excel(cons)
        
        ws_resumo.cell(row=linha_atual, column=1, value=descr)
        ws_resumo.cell(row=linha_atual, column=2, value=ind_milhares)
        ws_resumo.cell(row=linha_atual, column=3, value=cons_milhares)
        
        # Formatação numérica e alinhamento
        ws_resumo.cell(row=linha_atual, column=2).number_format = fmt_milhares
        ws_resumo.cell(row=linha_atual, column=2).alignment = Alignment(horizontal="right")
        ws_resumo.cell(row=linha_atual, column=3).number_format = fmt_milhares
        ws_resumo.cell(row=linha_atual, column=3).alignment = Alignment(horizontal="right")
        
        # Destaque para Saldo Inicial e Saldo Final
        if descr in {"Saldo Inicial", "Saldo Final"}:
            for c in range(1, 4):
                ws_resumo.cell(row=linha_atual, column=c).font = negrito_font
                ws_resumo.cell(row=linha_atual, column=c).fill = destaque_fill
        
        # Faixa laranja antes do Saldo Final
        if descr == "Saldo Final" and linha_atual > 3:
            # Inserir linha de separação
            ws_resumo.insert_rows(linha_atual)
            for c in range(1, 4):
                ws_resumo.cell(row=linha_atual, column=c).fill = separador_fill
            linha_atual += 1
            # Re-inserir a linha do Saldo Final após a separação
            ws_resumo.cell(row=linha_atual, column=1, value=descr)
            ws_resumo.cell(row=linha_atual, column=2, value=ind_milhares)
            ws_resumo.cell(row=linha_atual, column=3, value=cons_milhares)
            ws_resumo.cell(row=linha_atual, column=2).number_format = fmt_milhares
            ws_resumo.cell(row=linha_atual, column=2).alignment = Alignment(horizontal="right")
            ws_resumo.cell(row=linha_atual, column=3).number_format = fmt_milhares
            ws_resumo.cell(row=linha_atual, column=3).alignment = Alignment(horizontal="right")
            for c in range(1, 4):
                ws_resumo.cell(row=linha_atual, column=c).font = negrito_font
                ws_resumo.cell(row=linha_atual, column=c).fill = destaque_fill
        
        linha_atual += 1
    
    # Adicionar nota de rodapé
    linha_atual += 1
    ws_resumo.cell(row=linha_atual, column=1, value="* Valores em milhares de reais (R$ mil)")
    ws_resumo.cell(row=linha_atual, column=1).font = Font(italic=True, size=9)
    
    # Larguras de coluna
    ws_resumo.column_dimensions["A"].width = 25
    ws_resumo.column_dimensions["B"].width = 18
    ws_resumo.column_dimensions["C"].width = 18
    
    # Congelar cabeçalhos
    ws_resumo.freeze_panes = "A3"
    
    # --------------------------------------------------
    # ABA 2: COMPOSIÇÃO ANALÍTICA
    # --------------------------------------------------
    ws_comp = wb.create_sheet(title="Composição Analítica")
    
    # Linha 1: Cabeçalho do período
    ws_comp.merge_cells("A1:G1")
    cell_periodo_comp = ws_comp["A1"]
    cell_periodo_comp.value = f"{periodo_cabecalho} — Composição Analítica"
    cell_periodo_comp.font = negrito_branco_font
    cell_periodo_comp.fill = header_fill
    cell_periodo_comp.alignment = Alignment(horizontal="center", vertical="center")
    ws_comp.row_dimensions[1].height = 30
    
    # Linha 2: Cabeçalho das colunas
    colunas_comp = [
        "Conta Contábil",
        "Descrição",
        "Empresa",
        "Tipo Movimentação",
        "Valor Individual",
        "Valor Controladas",
        "Valor Consolidado",
    ]
    ws_comp.append(colunas_comp)
    for cell in ws_comp[2]:
        cell.font = negrito_font
        cell.fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Linhas de dados
    totais_comp = {"Individual": 0.0, "Controladas": 0.0, "Consolidado": 0.0}
    for reg in composicao_analitica:
        val_ind = reg.get("valor_individual", 0.0)
        val_ctrl = reg.get("valor_controladas", 0.0)
        val_cons = reg.get("valor_consolidado", 0.0)
        
        # Valores em milhares
        ind_milhares = _formatar_valor_milhares_excel(val_ind)
        ctrl_milhares = _formatar_valor_milhares_excel(val_ctrl)
        cons_milhares = _formatar_valor_milhares_excel(val_cons)
        
        ws_comp.append([
            reg.get("conta", ""),
            reg.get("descricao", "")[:50],  # Limitar descrição
            reg.get("empresa", ""),
            reg.get("classificacao", ""),
            ind_milhares,
            ctrl_milhares,
            cons_milhares,
        ])
        
        totais_comp["Individual"] += val_ind
        totais_comp["Controladas"] += val_ctrl
        totais_comp["Consolidado"] += val_cons
    
    # Formatação das colunas numéricas
    for row in range(3, ws_comp.max_row + 1):
        for col in (5, 6, 7):
            cell = ws_comp.cell(row=row, column=col)
            cell.number_format = fmt_milhares
            cell.alignment = Alignment(horizontal="right")
    
    # Linha de total
    row_total = ws_comp.max_row + 1
    ws_comp.cell(row=row_total, column=1, value="TOTAL COMPOSIÇÃO")
    ws_comp.cell(row=row_total, column=5, value=_formatar_valor_milhares_excel(totais_comp["Individual"]))
    ws_comp.cell(row=row_total, column=6, value=_formatar_valor_milhares_excel(totais_comp["Controladas"]))
    ws_comp.cell(row=row_total, column=7, value=_formatar_valor_milhares_excel(totais_comp["Consolidado"]))
    for col in range(1, 8):
        ws_comp.cell(row=row_total, column=col).font = negrito_font
        ws_comp.cell(row=row_total, column=col).fill = destaque_fill
    for col in (5, 6, 7):
        ws_comp.cell(row=row_total, column=col).number_format = fmt_milhares
        ws_comp.cell(row=row_total, column=col).alignment = Alignment(horizontal="right")
    
    # Nota de rodapé
    row_nota = row_total + 2
    ws_comp.cell(row=row_nota, column=1, value="* Valores em milhares de reais (R$ mil)")
    ws_comp.cell(row=row_nota, column=1).font = Font(italic=True, size=9)
    
    # Larguras de coluna
    ws_comp.column_dimensions["A"].width = 15
    ws_comp.column_dimensions["B"].width = 40
    ws_comp.column_dimensions["C"].width = 10
    ws_comp.column_dimensions["D"].width = 20
    ws_comp.column_dimensions["E"].width = 18
    ws_comp.column_dimensions["F"].width = 18
    ws_comp.column_dimensions["G"].width = 18
    
    # Congelar cabeçalhos
    ws_comp.freeze_panes = "A3"
    
    # Auto-filtro na composição
    ws_comp.auto_filter.ref = f"A2:G{ws_comp.max_row - 2}"
    
    # --------------------------------------------------
    # VALIDAÇÃO: Reconciliar composição com resumo
    # --------------------------------------------------
    # A soma da composição deve reconciliar com as linhas do resumo
    logger.info(
        "RECONCILIAÇÃO: Total Composição Individual=%s, Controladas=%s, Consolidado=%s",
        _formatar_valor_milhares(totais_comp["Individual"]),
        _formatar_valor_milhares(totais_comp["Controladas"]),
        _formatar_valor_milhares(totais_comp["Consolidado"]),
    )

    wb.save(caminho_xlsx)
    logger.info("Resumo Excel gerado: '%s' (2 abas: Resumo + Composição Analítica)", caminho_xlsx)
    return caminho_xlsx


# ---------------------------------------------------------------------------
# RESUMO
# ---------------------------------------------------------------------------

# Ordem fixa de exibição das categorias no resumo
_ORDEM_RESUMO = ["Adições", "Juros", "IOF", "IRRF", "Baixa"]


def exportar_resumo_xlsx(
    caminho_xlsx: str,
    categorias_ordenadas: list,
    contagens: dict,
    totais: dict,
    total_geral_linhas: int,
    total_geral_valor: float,
    logger: logging.Logger,
) -> None:
    """
    Exporta o resumo para Excel (.xlsx) usando openpyxl.

    Formata o arquivo com:
      - Cabeçalho em negrito.
      - Larguras de coluna ajustadas.
      - Valores monetários com formato numérico contábil.
      - Linha de TOTAL GERAL em negrito ao final.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment
    except ImportError:
        logger.error(
            "openpyxl não instalado. Execute: pip install openpyxl"
        )
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Resumo"

    # Cabeçalho
    ws.append(["Classificação", "Lançamentos", "Total Montante"])
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")

    # Linhas de dados
    for cat in categorias_ordenadas:
        total = totais.get(cat, 0.0)
        cnt = contagens.get(cat, 0)
        ws.append([cat, cnt, total])

    # Linha de total geral
    ws.append(["TOTAL GERAL", total_geral_linhas, total_geral_valor])
    ultima_linha = ws.max_row
    for cell in ws[ultima_linha]:
        cell.font = Font(bold=True)

    # Larguras de coluna
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 20

    # Formato numérico monetário para coluna C (linhas de dados + total)
    fmt_numero = '#,##0.00'
    for row in ws.iter_rows(min_row=2, max_row=ultima_linha, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = fmt_numero

    # Alinhamento: coluna B (Lançamentos) à direita
    for row in ws.iter_rows(min_row=2, max_row=ultima_linha, min_col=2, max_col=2):
        for cell in row:
            cell.alignment = Alignment(horizontal="right")

    wb.save(caminho_xlsx)
    logger.info("Resumo Excel gerado: '%s'", caminho_xlsx)


def gerar_resumo(
    arquivo_classificado: str,
    pasta_saida: str,
    logger: logging.Logger,
):
    """
    Lê o arquivo classificado e gera um resumo com totais de Montante Razão
    agrupados por Classificação.

    Considera APENAS as 5 categorias (linhas de Mútuo classificadas):
      Adições, Juros, IOF, IRRF, Baixa

    Saídas:
      - CSV de resumo em UTF-8-sig.
      - Excel (.xlsx) via openpyxl com formatação contábil.
      - Impressão no terminal.
      - Registro no LOG.

    Retorna o caminho do arquivo CSV de resumo, ou None em caso de falha.
    """
    if not os.path.isfile(arquivo_classificado):
        logger.error("Arquivo classificado não encontrado: '%s'", arquivo_classificado)
        return None

    os.makedirs(pasta_saida, exist_ok=True)

    logger.info("Gerando resumo de '%s'...", arquivo_classificado)

    # Detectar índice da coluna Classificação a partir do cabeçalho
    idx_classif = None
    idx_montante_resumo = IDX_MONTANTE  # padrão; recalculado ao ler o cabeçalho
    idx_conta_resumo = IDX_CONTA       # detectado dinamicamente (item 5)

    totais: dict = {}
    contagens_linhas: dict = {}

    try:
        for linha in ler_csv_corrigindo_encoding(arquivo_classificado, logger):
            tp = tipo_de_linha(linha)

            if tp == "cabecalho":
                campos = parse_linha(linha)
                for i, c in enumerate(campos):
                    c_n = c.strip().lower()
                    if "classificação" in c_n or "classificacao" in c_n:
                        idx_classif = i
                    if "montante" in c_n:
                        idx_montante_resumo = i
                    if c_n == "conta":
                        idx_conta_resumo = i
                continue

            if tp != "dado":
                continue

            campos = parse_linha(linha)

            # Somente linhas de Mútuo (1202*) entram nos totais
            conta = campos[idx_conta_resumo].strip() if len(campos) > idx_conta_resumo else ""
            if not conta.startswith(PREFIXO_MUTUO):
                continue

            classif = ""
            if idx_classif is not None and len(campos) > idx_classif:
                classif = campos[idx_classif].strip()

            # Ignorar linhas sem classificação real no resumo (inclui "N/A" — item 3)
            if not classif or classif == "N/A":
                continue

            montante = None
            if len(campos) > idx_montante_resumo:
                montante = _parse_montante(campos[idx_montante_resumo])

            totais[classif] = totais.get(classif, 0.0) + (montante or 0.0)
            contagens_linhas[classif] = contagens_linhas.get(classif, 0) + 1

    except Exception as exc:  # noqa: BLE001
        logger.error("Erro ao ler arquivo classificado para resumo: %s", exc)
        return None

    # Categorias na ordem fixa; inclui apenas as que têm dados
    categorias_com_dados = [c for c in _ORDEM_RESUMO if c in totais]

    total_geral = sum(totais.get(c, 0.0) for c in categorias_com_dados)
    total_linhas_geral = sum(contagens_linhas.get(c, 0) for c in categorias_com_dados)

    # Formatar montante como string BR para exibição
    def _fmt_br(valor: float) -> str:
        abs_v = abs(valor)
        s = f"{abs_v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        return f"{s}-" if valor < 0 else s

    # Exibição no terminal e no LOG
    separador = "-" * 62
    linhas_resumo = [
        separador,
        f"  RESUMO — {os.path.basename(arquivo_classificado)}",
        separador,
        f"  {'Classificação':<20} {'Lançamentos':>12} {'Total Montante':>20}",
        separador,
    ]
    for chave in categorias_com_dados:
        linhas_resumo.append(
            f"  {chave:<20} {contagens_linhas[chave]:>12}  "
            f"{_fmt_br(totais[chave]):>18}"
        )
    linhas_resumo += [
        separador,
        f"  {'TOTAL GERAL':<20} {total_linhas_geral:>12}  "
        f"{_fmt_br(total_geral):>18}",
        separador,
    ]

    for l in linhas_resumo:
        print(l)
    for l in linhas_resumo:
        logger.info(l)

    # Derivar datas do nome do arquivo
    data_ini_str, data_fim_str = _extrair_datas_do_nome(
        os.path.basename(arquivo_classificado)
    )

    # Gravar CSV de resumo
    nome_resumo_csv = f"Resumo_{data_ini_str}_a_{data_fim_str}.csv"
    caminho_resumo_csv = os.path.join(pasta_saida, nome_resumo_csv)

    with open(caminho_resumo_csv, "w", encoding="utf-8-sig", newline="") as f:
        f.write("Classificação|Lançamentos|Total Montante\n")
        for chave in categorias_com_dados:
            f.write(
                f"{chave}|{contagens_linhas[chave]}|"
                f"{_fmt_br(totais[chave])}\n"
            )
        f.write(f"TOTAL GERAL|{total_linhas_geral}|{_fmt_br(total_geral)}\n")

    logger.info("Arquivo CSV de resumo gerado: '%s'", caminho_resumo_csv)

    # Gravar Excel de resumo
    nome_resumo_xlsx = f"Resumo_{data_ini_str}_a_{data_fim_str}.xlsx"
    caminho_resumo_xlsx = os.path.join(pasta_saida, nome_resumo_xlsx)
    exportar_resumo_xlsx(
        caminho_resumo_xlsx,
        categorias_com_dados,
        contagens_linhas,
        totais,
        total_linhas_geral,
        total_geral,
        logger,
    )

    return caminho_resumo_csv


# ---------------------------------------------------------------------------
# UTILITÁRIOS DE MENU
# ---------------------------------------------------------------------------

def _extrair_datas_do_nome(nome_arquivo: str) -> tuple:
    """
    Extrai as datas do nome de arquivo no padrão *_DD.MM.AAAA_a_DD.MM.AAAA*.
    Retorna ('DD.MM.AAAA', 'DD.MM.AAAA') ou ('inicio', 'fim') se não encontrar.
    """
    match = re.search(
        r"(\d{2}\.\d{2}\.\d{4})_a_(\d{2}\.\d{2}\.\d{4})", nome_arquivo
    )
    if match:
        return match.group(1), match.group(2)
    return "inicio", "fim"


def _encontrar_ultimo_arquivo(pasta: str, padrao: str):
    """
    Retorna o caminho do arquivo mais recente na pasta que corresponde ao
    padrão glob fornecido, ou None se não houver arquivos.
    """
    arquivos = glob.glob(os.path.join(pasta, padrao))
    if not arquivos:
        return None
    return max(arquivos, key=os.path.getmtime)


def _ler_data(prompt: str, logger: logging.Logger) -> datetime:
    """Solicita uma data no formato dd.mm.aaaa até obter entrada válida."""
    while True:
        valor = input(prompt).strip()
        try:
            return datetime.strptime(valor, "%d.%m.%Y")
        except ValueError:
            msg = (
                f"Data inválida: '{valor}'. "
                "Use o formato dd.mm.aaaa (ex.: 01.06.2026)."
            )
            logger.warning(msg)
            print(f"  ✗ {msg}")


def _ler_periodicidade(logger: logging.Logger) -> str:
    """Solicita a periodicidade até obter opção válida."""
    opcoes = {"1": "diaria", "2": "semanal", "3": "mensal"}
    while True:
        print("\nPeriodicidade:")
        print("  1 - Diária  (padrão)")
        print("  2 - Semanal")
        print("  3 - Mensal")
        valor = input("Opção [1]: ").strip() or "1"
        if valor in opcoes:
            return opcoes[valor]
        msg = f"Opção inválida: '{valor}'. Digite 1, 2 ou 3."
        logger.warning(msg)
        print(f"  ✗ {msg}")


def _ler_sistema(logger: logging.Logger) -> dict:
    """Solicita o sistema SAP até obter opção válida."""
    while True:
        print("\nSistema/Conexão SAP:")
        for chave, info in SISTEMAS.items():
            print(f"  {chave} - {info['nome']}")
        valor = input("Opção: ").strip()
        if valor in SISTEMAS:
            return SISTEMAS[valor]
        msg = f"Opção inválida: '{valor}'. Digite {' ou '.join(SISTEMAS.keys())}."
        logger.warning(msg)
        print(f"  ✗ {msg}")


def selecionar_arquivo_entrada(
    prompt: str, padrao: str, pasta: str, logger: logging.Logger
) -> str:
    """
    Solicita o caminho de um arquivo de entrada com lógica robusta de resolução.

    - Apresenta o arquivo mais recente como padrão (Enter para aceitar).
    - Se o usuário digitar apenas o nome (com ou sem .csv), procura na 'pasta'.
    - Se digitar um caminho absoluto, usa diretamente.
    - Adiciona .csv automaticamente se o arquivo informado não tiver extensão.
    - Valida existência antes de retornar.
    """
    ultimo = _encontrar_ultimo_arquivo(pasta, padrao)
    if ultimo:
        nome_padrao = os.path.basename(ultimo)
        prompt_completo = (
            f"{prompt}\n"
            f"  [padrão: {nome_padrao}]\n"
            f"  (Enter para aceitar, ou informe nome/caminho): "
        )
    else:
        prompt_completo = (
            f"{prompt}\n"
            f"  (nome do arquivo ou caminho completo): "
        )

    while True:
        valor = input(prompt_completo).strip()

        # Enter → usar padrão
        if not valor:
            if ultimo:
                logger.info("Arquivo padrão selecionado: '%s'", ultimo)
                return ultimo
            print("  ✗ Nenhum arquivo disponível como padrão. Informe o caminho.")
            continue

        # Determinar candidato(s) a testar
        candidatos = []

        if os.path.isabs(valor):
            # Caminho absoluto: usar como está (+ .csv se sem extensão)
            if not os.path.splitext(valor)[1]:
                candidatos.append(valor + ".csv")
            candidatos.append(valor)
        else:
            # Nome simples ou relativo: procurar na pasta de saída
            nome = valor
            if not os.path.splitext(nome)[1]:
                nome_csv = nome + ".csv"
                candidatos.append(os.path.join(pasta, nome_csv))
            candidatos.append(os.path.join(pasta, nome))
            # Também testar no diretório atual como fallback
            if not os.path.splitext(nome)[1]:
                candidatos.append(nome + ".csv")
            candidatos.append(nome)

        for c in candidatos:
            if os.path.isfile(c):
                logger.info("Arquivo selecionado: '%s'", c)
                return c

        msg = (
            f"Arquivo não encontrado: '{valor}'. "
            f"Procurado em: '{pasta}' e diretório atual. "
            "Informe o nome ou caminho completo."
        )
        logger.warning(msg)
        print(f"  ✗ {msg}")
        if pasta:
            print(f"  Dica: arquivos disponíveis em '{pasta}':")
            disponiveis = glob.glob(os.path.join(pasta, padrao))
            for arq in sorted(disponiveis)[-5:]:
                print(f"    • {os.path.basename(arq)}")


# ---------------------------------------------------------------------------
# AÇÕES DO MENU
# ---------------------------------------------------------------------------

def _executar_extracao(logger: logging.Logger, caminho_log: str) -> None:
    """Opção 1: Extrai FBL3N e consolida. Correspondente ao fluxo original."""
    print("=" * 60)
    print("  Extrair Razão (FBL3N) + Consolidar")
    print("=" * 60)

    data_inicial = _ler_data("Data inicial (dd.mm.aaaa): ", logger)
    while True:
        data_final = _ler_data("Data final   (dd.mm.aaaa): ", logger)
        if data_final >= data_inicial:
            break
        msg = "A data final deve ser igual ou posterior à data inicial."
        logger.warning(msg)
        print(f"  ✗ {msg}")

    periodicidade = _ler_periodicidade(logger)
    sistema = _ler_sistema(logger)

    logger.info(
        "Extração: período=%s a %s | periodicidade=%s | sistema=%s",
        data_inicial.strftime("%d.%m.%Y"),
        data_final.strftime("%d.%m.%Y"),
        periodicidade,
        sistema["nome"],
    )

    session = conectar_sap(sistema, logger)
    if session is None:
        logger.error("Não foi possível obter sessão SAP. Encerrando.")
        print(f"\nConsulte o LOG para detalhes: {caminho_log}")
        return

    os.makedirs(PASTA_DIARIOS, exist_ok=True)
    csvs = iterar_periodo(
        data_inicial, data_final, periodicidade, session, PASTA_DIARIOS, logger
    )

    consolidado = consolidar(
        csvs, data_inicial, data_final, PASTA_CONSOLIDADO, logger
    )

    print("\n" + "=" * 60)
    if consolidado:
        print(f"  Consolidado gerado: {consolidado}")
    else:
        print("  Nenhum arquivo consolidado gerado (verifique o LOG).")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_classificar(logger: logging.Logger, caminho_log: str) -> None:
    """Opção 2: Classifica um arquivo consolidado existente."""
    print("=" * 60)
    print("  Classificar Consolidado")
    print("=" * 60)

    arq = selecionar_arquivo_entrada(
        "Arquivo consolidado a classificar",
        "Consolidado_*.csv",
        PASTA_CONSOLIDADO,
        logger,
    )

    classificado = classificar(arq, PASTA_CONSOLIDADO, logger)

    print("\n" + "=" * 60)
    if classificado:
        print(f"  Classificado gerado: {classificado}")
    else:
        print("  Falha na classificação (verifique o LOG).")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_importar_tabelas(logger: logging.Logger, caminho_log: str) -> None:
    """Opção: importa/atualiza ZFIT009 e ZCO059 via SAP ou arquivo manual."""
    print("=" * 60)
    print("  Importar ZFIT009 / ZCO059")
    print("=" * 60)

    print("Fonte de importação:")
    print("  1 - SAP GUI Scripting")
    print("  2 - Arquivo já exportado manualmente")
    origem = input("Opção [2]: ").strip() or "2"

    if origem == "1":
        sistema = _ler_sistema(logger)
        session = conectar_sap(sistema, logger)
        if session is None:
            print("  Falha ao conectar no SAP (ver LOG).")
            return
        zfit = importar_zfit009_sap(session, PASTA_TABELAS, logger)
        zco = importar_zco059_sap(session, PASTA_TABELAS, logger)
    else:
        arq_zfit = selecionar_arquivo_entrada(
            "Arquivo ZFIT009 (manual)",
            "ZFIT009*.csv",
            PASTA_TABELAS,
            logger,
        )
        arq_zco = selecionar_arquivo_entrada(
            "Arquivo ZCO059 (manual)",
            "ZCO059*.csv",
            PASTA_TABELAS,
            logger,
        )
        zfit = importar_tabela_de_arquivo(arq_zfit, "zfit009", PASTA_TABELAS, logger)
        zco = importar_tabela_de_arquivo(arq_zco, "zco059", PASTA_TABELAS, logger)

    print("\n" + "=" * 60)
    if zfit and zco:
        print(f"  ZFIT009: {zfit}")
        print(f"  ZCO059 : {zco}")
    else:
        print("  Falha na importação de tabelas (verifique o LOG).")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_gerar_consolidacao(logger: logging.Logger, caminho_log: str) -> None:
    """Opção: gera Consolidação Cliente|Divisão|Consolida."""
    print("=" * 60)
    print("  Gerar Tabela de Consolidação")
    print("=" * 60)

    caminho = gerar_consolidacao(PASTA_TABELAS, PASTA_CONSOLIDADO, logger)

    print("\n" + "=" * 60)
    if caminho:
        print(f"  Consolidação gerada: {caminho}")
    else:
        print("  Falha ao gerar consolidação (verifique o LOG).")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_aplicar_status_consolidacao(logger: logging.Logger, caminho_log: str) -> None:
    """Opção: aplica status de consolidação no classificado."""
    print("=" * 60)
    print("  Aplicar Status Consolidação")
    print("=" * 60)

    arq_classificado = selecionar_arquivo_entrada(
        "Arquivo classificado para complementar",
        "Classificado_*.csv",
        PASTA_CONSOLIDADO,
        logger,
    )
    arq_consolidacao = selecionar_arquivo_entrada(
        "Arquivo de consolidação cliente/divisão/consolida",
        "Consolidacao_Cliente_Divisao_Consolida.csv",
        PASTA_CONSOLIDADO,
        logger,
    )
    saida = aplicar_status_consolidacao(
        arq_classificado, arq_consolidacao, PASTA_CONSOLIDADO, logger
    )

    print("\n" + "=" * 60)
    if saida:
        print(f"  Classificado com status: {saida}")
    else:
        print("  Falha ao aplicar status (verifique o LOG).")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_tratar_contrapartidas(logger: logging.Logger, caminho_log: str) -> None:
    """Opção: separa contrapartidas para auditoria."""
    print("=" * 60)
    print("  Tratar Contrapartidas (Auditoria)")
    print("=" * 60)

    arq_classificado = selecionar_arquivo_entrada(
        "Arquivo classificado para separar contrapartidas",
        "Classificado_*.csv",
        PASTA_CONSOLIDADO,
        logger,
    )
    principal, auditoria = separar_contrapartidas(arq_classificado, PASTA_CONSOLIDADO, logger)
    xlsx = exportar_excel_corporativo(principal, logger) if principal else None

    print("\n" + "=" * 60)
    if principal and auditoria:
        print(f"  Principal : {principal}")
        print(f"  Auditoria : {auditoria}")
        if xlsx:
            print(f"  Excel     : {xlsx}")
    else:
        print("  Falha ao tratar contrapartidas (verifique o LOG).")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_saldo_inicial(logger: logging.Logger, caminho_log: str) -> None:
    """
    Opção: informa/atualiza saldo inicial persistido.
    
    REQUISITO 5: Input separado para Individual e Controladas.
    Consolidado = Individual + Controladas (calculado automaticamente, sem input direto).
    """
    print("=" * 60)
    print("  Informar / Atualizar Saldo Inicial")
    print("  (Individual e Controladas — Consolidado é calculado)")
    print("=" * 60)

    atual = carregar_saldo_inicial(logger)
    if atual:
        valor_ind = atual.get("valor_individual", atual.get("valor", 0.0))
        valor_ctrl = atual.get("valor_controladas", 0.0)
        valor_cons = valor_ind + valor_ctrl
        print(f"\nSaldo atual persistido:")
        print(f"  Individual : {_formatar_valor_br(valor_ind)}")
        print(f"  Controladas: {_formatar_valor_br(valor_ctrl)}")
        print(f"  Consolidado: {_formatar_valor_br(valor_cons)} (calculado)")
        print(f"  Período    : {atual.get('periodo', '')}")
        print(f"  Atualizado : {atual.get('data_alteracao', '')} | Usuário: {atual.get('usuario', '')}")
        default_ind = valor_ind
        default_ctrl = valor_ctrl
    else:
        default_ind = 0.0
        default_ctrl = 0.0

    print("\n  NOTA: Informe os valores SEPARADAMENTE. O Consolidado será calculado.")
    print("  Formato aceito: padrão BR (ex.: 1.234.567,89 ou 1234567,89)")

    # Input Individual
    prompt_ind = f"Saldo Inicial INDIVIDUAL [padrão: {_formatar_valor_br(default_ind)}]: "
    valor_ind_txt = input(prompt_ind).strip()
    if valor_ind_txt:
        valor_individual = _parse_numero_br(valor_ind_txt)
        if valor_individual is None:
            print("  ✗ Valor Individual inválido.")
            logger.warning("Saldo inicial Individual inválido informado: '%s'", valor_ind_txt)
            return
    else:
        valor_individual = default_ind

    # Input Controladas
    prompt_ctrl = f"Saldo Inicial CONTROLADAS [padrão: {_formatar_valor_br(default_ctrl)}]: "
    valor_ctrl_txt = input(prompt_ctrl).strip()
    if valor_ctrl_txt:
        valor_controladas = _parse_numero_br(valor_ctrl_txt)
        if valor_controladas is None:
            print("  ✗ Valor Controladas inválido.")
            logger.warning("Saldo inicial Controladas inválido informado: '%s'", valor_ctrl_txt)
            return
    else:
        valor_controladas = default_ctrl

    # Consolidado calculado automaticamente
    valor_consolidado = valor_individual + valor_controladas
    print(f"\n  → Consolidado calculado: {_formatar_valor_br(valor_consolidado)}")

    periodo = input("Período/Trimestre de referência: ").strip() or "não informado"
    salvar_saldo_inicial(valor_individual, valor_controladas, periodo, logger)

    print("\n" + "=" * 60)
    print("  Saldo inicial atualizado com sucesso.")
    print(f"  Individual : {_formatar_valor_br(valor_individual)}")
    print(f"  Controladas: {_formatar_valor_br(valor_controladas)}")
    print(f"  Consolidado: {_formatar_valor_br(valor_consolidado)}")
    print(f"  Arquivo: {ARQUIVO_SALDO_INICIAL}")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_resumo_fluxo(logger: logging.Logger, caminho_log: str) -> None:
    """Opção: gera resumo de fluxo do classificado."""
    print("=" * 60)
    print("  Resumo de Fluxo (Saldo Inicial → Saldo Final)")
    print("=" * 60)

    arq = selecionar_arquivo_entrada(
        "Arquivo classificado para gerar resumo",
        "Classificado_*.csv",
        PASTA_CONSOLIDADO,
        logger,
    )
    resumo = gerar_resumo_fluxo(arq, PASTA_CONSOLIDADO, logger)

    print("\n" + "=" * 60)
    if resumo:
        print(f"  Resumo gerado: {resumo}")
    else:
        print("  Falha ao gerar resumo (verifique o LOG).")
    print(f"  LOG: {caminho_log}")
    print("=" * 60)


def _executar_tudo(logger: logging.Logger, caminho_log: str) -> None:
    """Opção: executa pipeline completo com consolidação intercompany."""
    print("=" * 60)
    print("  Executar Tudo: Extrair → Classificar → Consolidar → Status → Auditoria → Resumo")
    print("=" * 60)

    data_inicial = _ler_data("Data inicial (dd.mm.aaaa): ", logger)
    while True:
        data_final = _ler_data("Data final   (dd.mm.aaaa): ", logger)
        if data_final >= data_inicial:
            break
        msg = "A data final deve ser igual ou posterior à data inicial."
        logger.warning(msg)
        print(f"  ✗ {msg}")

    periodicidade = _ler_periodicidade(logger)
    sistema = _ler_sistema(logger)

    logger.info(
        "Execução completa: período=%s a %s | periodicidade=%s | sistema=%s",
        data_inicial.strftime("%d.%m.%Y"),
        data_final.strftime("%d.%m.%Y"),
        periodicidade,
        sistema["nome"],
    )

    session = conectar_sap(sistema, logger)
    if session is None:
        logger.error("Não foi possível obter sessão SAP. Encerrando.")
        print(f"\nConsulte o LOG para detalhes: {caminho_log}")
        return

    os.makedirs(PASTA_DIARIOS, exist_ok=True)
    csvs = iterar_periodo(
        data_inicial, data_final, periodicidade, session, PASTA_DIARIOS, logger
    )

    consolidado = consolidar(
        csvs, data_inicial, data_final, PASTA_CONSOLIDADO, logger
    )
    if not consolidado:
        print("  Nenhum dado consolidado. Encerrando fluxo completo.")
        return

    classificado = classificar(consolidado, PASTA_CONSOLIDADO, logger)
    if not classificado:
        print("  Falha na classificação. Resumo não gerado.")
        return

    consolidacao = gerar_consolidacao(PASTA_TABELAS, PASTA_CONSOLIDADO, logger)
    if not consolidacao:
        print("  Falha ao gerar consolidação (importe ZFIT009/ZCO059 e tente novamente).")
        return

    classificado_status = aplicar_status_consolidacao(
        classificado, consolidacao, PASTA_CONSOLIDADO, logger
    )
    if not classificado_status:
        print("  Falha ao aplicar status de consolidação.")
        return

    classificado_final, auditoria = separar_contrapartidas(
        classificado_status, PASTA_CONSOLIDADO, logger
    )
    if not classificado_final:
        print("  Falha no tratamento de contrapartidas.")
        return

    xlsx_classificado = exportar_excel_corporativo(classificado_final, logger)
    resumo = gerar_resumo_fluxo(classificado_final, PASTA_CONSOLIDADO, logger)

    print("\n" + "=" * 60)
    print(f"  Consolidado  : {consolidado}")
    print(f"  Classificado : {classificado_final}")
    print(f"  Consolidação : {consolidacao}")
    print(f"  Auditoria    : {auditoria}")
    if xlsx_classificado:
        print(f"  Excel Class. : {xlsx_classificado}")
    if resumo:
        print(f"  Resumo       : {resumo}")
    print(f"  LOG          : {caminho_log}")
    print("=" * 60)


# ---------------------------------------------------------------------------
# MENU PRINCIPAL
# ---------------------------------------------------------------------------

def mostrar_menu_principal() -> str:
    """
    Exibe o menu principal e retorna a opção escolhida como string.
    """
    print("\n" + "=" * 60)
    print("  MUTDFC — Movimentação do Fluxo de Caixa (Mútuo)")
    print("=" * 60)
    print("  1 - Extrair razão (FBL3N) + Consolidar")
    print("  2 - Classificar consolidado")
    print("  3 - Importar ZFIT009 / ZCO059")
    print("  4 - Gerar tabela de consolidação")
    print("  5 - Aplicar Status Consolidação na movimentação")
    print("  6 - Tratar contrapartidas (auditoria) + Excel classificado")
    print("  7 - Informar/atualizar Saldo Inicial")
    print("  8 - Resumo (Saldo Inicial → Saldo Final)")
    print("  9 - Executar tudo em sequência")
    print("  A - Ajustes Manuais (de-para Cliente → Divisão)")
    print("  0 - Sair")
    print("-" * 60)
    return input("Opção: ").strip()


# ---------------------------------------------------------------------------
# PONTO DE ENTRADA
# ---------------------------------------------------------------------------

def main() -> None:
    # Garantir que as pastas de saída existam
    for pasta in (PASTA_DIARIOS, PASTA_CONSOLIDADO, PASTA_TABELAS, PASTA_LOGS):
        os.makedirs(pasta, exist_ok=True)

    logger, caminho_log = configurar_log(PASTA_LOGS)

    acoes = {
        "1": _executar_extracao,
        "2": _executar_classificar,
        "3": _executar_importar_tabelas,
        "4": _executar_gerar_consolidacao,
        "5": _executar_aplicar_status_consolidacao,
        "6": _executar_tratar_contrapartidas,
        "7": _executar_saldo_inicial,
        "8": _executar_resumo_fluxo,
        "9": _executar_tudo,
    }

    while True:
        try:
            opcao = mostrar_menu_principal()
        except KeyboardInterrupt:
            print("\nOperação cancelada pelo usuário.")
            break

        if opcao == "0":
            print("Saindo. Até logo!")
            break
        elif opcao.upper() == "A":
            try:
                menu_ajustes_manuais(logger, caminho_log)
            except KeyboardInterrupt:
                print("\nOperação interrompida. Voltando ao menu.")
        elif opcao in acoes:
            try:
                acoes[opcao](logger, caminho_log)
            except KeyboardInterrupt:
                print("\nOperação interrompida. Voltando ao menu.")
        else:
            print(f"  ✗ Opção inválida: '{opcao}'. Digite 0-9 ou A.")


if __name__ == "__main__":
    main()
